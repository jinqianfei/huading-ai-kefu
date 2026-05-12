#!/usr/bin/env python3
"""
生成华鼎标准出库单
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# 华鼎模板字段定义（35列）
HUADING_FIELDS = [
    ("序号", "A"),
    ("门店编号", "B"),
    ("门店三方编码", "C"),
    ("仓库编码", "D"),
    ("加急程度", "E"),
    ("商品编码", "F"),
    ("商品名称", "G"),
    ("规格", "H"),
    ("数量", "I"),
    ("单位", "J"),
    ("商品SKU编号", "K"),
    ("商品三方SPEC编号", "L"),
    ("单位类型", "M"),
    ("指定库存状态", "N"),
    ("出库类型", "O"),
    ("配送方式", "P"),
    ("指定车型", "Q"),
    ("是否垫付", "R"),
    ("付款方式", "S"),
    ("快递公司", "T"),
    ("单价", "U"),
    ("总金额", "V"),
    ("是否制定批次", "W"),
    ("批次号", "X"),
    ("生产日期", "Y"),
    ("生产厂家编号", "Z"),
    ("门店收货地址编码", "AA"),
    ("三方单号", "AB"),
    ("业务模式", "AC"),
    ("业务类型", "AD"),
    ("收货人", "AE"),
    ("联系电话", "AF"),
    ("收货地址", "AG"),
    ("C端快递公司", "AH"),
    ("备注", "AI"),
]

# 默认值字段
DEFAULT_VALUES = {
    "加急程度": "0",
    "指定库存状态": "正常",
    "出库类型": "201",
    "是否垫付": "否",
    "付款方式": "预付",
    "是否制定批次": "否",
    "业务模式": "B2C",
    "仓库编码": "5",
}

def create_styles(wb):
    """创建表头和数据样式"""
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    header_style = wb.add_style({
        "font": header_font,
        "fill": header_fill,
        "alignment": header_alignment,
        "border": Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
    })
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_style = wb.add_style({
        "alignment": data_alignment,
        "border": Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
    })
    
    return header_style, data_style

def set_column_widths(ws):
    """设置列宽"""
    widths = {
        "A": 8, "B": 15, "C": 15, "D": 12, "E": 10,
        "F": 15, "G": 25, "H": 15, "I": 10, "J": 10,
        "K": 18, "L": 20, "M": 12, "N": 14, "O": 12,
        "P": 12, "Q": 15, "R": 12, "S": 12, "T": 15,
        "U": 12, "V": 12, "W": 14, "X": 15, "Y": 14,
        "Z": 15, "AA": 18, "AB": 30, "AC": 12, "AD": 12,
        "AE": 12, "AF": 15, "AG": 35, "AH": 15, "AI": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def generate_third_party_order_no(customer_id, order_no):
    """生成三方单号: CUS_{cid}_{order}_{timestamp}"""
    ts = int(datetime.now().timestamp() * 1000)
    clean_cid = customer_id.replace(" ", "_")
    clean_order = order_no.replace(" ", "_")
    return f"CUS_{clean_cid}_{clean_order}_{ts}"

def main():
    # 订单数据（从之前解析得到）
    order_header = {
        "order_no": "DH-O-20260423-283742",
        "customer_name": "任丘北辛庄店",
        "contact_person": "肖文月",
        "contact_phone": "13171700566",
        "delivery_address": "河北省沧州市任丘市北辛庄镇北辛庄村制茶青年",
        "delivery_date": "2026-04-23",
    }
    
    # 商品明细
    order_items = [
        {"sku_code": "P1556689351", "sku_name": "1000塑杯", "spec": "规格：600个", "quantity": 1, "unit": "件"},
        {"sku_code": "P969550376", "sku_name": "速冻桑葚/新", "spec": "件：1kg*10袋/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P6467042236", "sku_name": "果糖-", "spec": "件：6.25kg*4", "quantity": 3, "unit": "件"},
        {"sku_code": "P5469555375", "sku_name": "制茶基底乳", "spec": "件：1L*12盒/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P8956720518", "sku_name": "果蜜-", "spec": "规格：6.25kg*4/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P3056699182", "sku_name": "凤梨味果酱", "spec": "规格：2kg*6瓶/件", "quantity": 3, "unit": "件"},
        {"sku_code": "P6856694304", "sku_name": "红豆罐头", "spec": "规格：12罐/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P8365640612", "sku_name": "原味晶球", "spec": "件：1kg*20袋", "quantity": 2, "unit": "件"},
        {"sku_code": "P9968023878", "sku_name": "三孔吸管-轻乳茶专用", "spec": "件：10包＊100支", "quantity": 1, "unit": "件"},
        {"sku_code": "P4869589451", "sku_name": "牛乳茶纸杯", "spec": "件：25只*20条/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P2969330615", "sku_name": "春香茉莉花茶（50袋*100g）", "spec": "件：100g×50袋", "quantity": 1, "unit": "件"},
        {"sku_code": "P5869725757", "sku_name": "圣代勺（1000支）", "spec": "件：1000支/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P7669860793", "sku_name": "奶冻粉蛋白固体饮料", "spec": "", "quantity": 1, "unit": "件"},
        {"sku_code": "P9056699290", "sku_name": "巧克力风味糖浆", "spec": "规格：6瓶/件", "quantity": 1, "unit": "件"},
        {"sku_code": "P3864643137", "sku_name": "制茶青年甄选纯牛奶", "spec": "件：1L*12瓶", "quantity": 3, "unit": "件"},
        {"sku_code": "P9264773821", "sku_name": "牛奶冰淇淋粉", "spec": "件：3kg*7袋", "quantity": 3, "unit": "件"},
        {"sku_code": "P9956693697", "sku_name": "抹茶冰淇淋粉", "spec": "规格：3kg*8袋", "quantity": 1, "unit": "箱"},
        {"sku_code": "P8269828408", "sku_name": "冷冻草莓浆", "spec": "件：1kg*12瓶", "quantity": 2, "unit": "件"},
        {"sku_code": "P1169843848", "sku_name": "冷冻蜜瓜浆", "spec": "件：1kg*12瓶", "quantity": 1, "unit": "件"},
    ]
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "华鼎标准出库单"
    
    # 设置列宽
    set_column_widths(ws)
    
    # 写入表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    
    # 表头行
    for field_name, col in HUADING_FIELDS:
        cell = ws[f"{col}1"]
        cell.value = field_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 生成三方单号
    third_party_order_no = generate_third_party_order_no("任丘北辛庄店", order_header["order_no"])
    
    # 写入数据行
    for row_idx, item in enumerate(order_items, start=2):
        # 序号
        ws[f"A{row_idx}"] = row_idx - 1
        # 门店编号 - 待确认
        ws[f"B{row_idx}"] = ""
        # 门店三方编码
        ws[f"C{row_idx}"] = ""
        # 仓库编码 - 默认值
        ws[f"D{row_idx}"] = DEFAULT_VALUES["仓库编码"]
        # 加急程度 - 默认值
        ws[f"E{row_idx}"] = DEFAULT_VALUES["加急程度"]
        # 商品编码
        ws[f"F{row_idx}"] = item["sku_code"]
        # 商品名称
        ws[f"G{row_idx}"] = item["sku_name"]
        # 规格
        ws[f"H{row_idx}"] = item["spec"]
        # 数量
        ws[f"I{row_idx}"] = item["quantity"]
        # 单位
        ws[f"J{row_idx}"] = item["unit"]
        # 商品SKU编号 - 与商品编码相同
        ws[f"K{row_idx}"] = item["sku_code"]
        # 商品三方SPEC编号
        ws[f"L{row_idx}"] = item["sku_code"]
        # 单位类型 - 待确认
        ws[f"M{row_idx}"] = ""
        # 指定库存状态 - 默认值
        ws[f"N{row_idx}"] = DEFAULT_VALUES["指定库存状态"]
        # 出库类型 - 默认值
        ws[f"O{row_idx}"] = DEFAULT_VALUES["出库类型"]
        # 配送方式 - 待确认
        ws[f"P{row_idx}"] = ""
        # 指定车型
        ws[f"Q{row_idx}"] = ""
        # 是否垫付 - 默认值
        ws[f"R{row_idx}"] = DEFAULT_VALUES["是否垫付"]
        # 付款方式 - 默认值
        ws[f"S{row_idx}"] = DEFAULT_VALUES["付款方式"]
        # 快递公司
        ws[f"T{row_idx}"] = ""
        # 单价
        ws[f"U{row_idx}"] = ""
        # 总金额
        ws[f"V{row_idx}"] = ""
        # 是否制定批次 - 默认值
        ws[f"W{row_idx}"] = DEFAULT_VALUES["是否制定批次"]
        # 批次号
        ws[f"X{row_idx}"] = ""
        # 生产日期
        ws[f"Y{row_idx}"] = ""
        # 生产厂家编号
        ws[f"Z{row_idx}"] = ""
        # 门店收货地址编码
        ws[f"AA{row_idx}"] = ""
        # 三方单号
        ws[f"AB{row_idx}"] = third_party_order_no
        # 业务模式 - 默认值
        ws[f"AC{row_idx}"] = DEFAULT_VALUES["业务模式"]
        # 业务类型 - 待确认
        ws[f"AD{row_idx}"] = ""
        # 收货人
        ws[f"AE{row_idx}"] = order_header["contact_person"]
        # 联系电话
        ws[f"AF{row_idx}"] = order_header["contact_phone"]
        # 收货地址
        ws[f"AG{row_idx}"] = order_header["delivery_address"]
        # C端快递公司
        ws[f"AH{row_idx}"] = ""
        # 备注
        ws[f"AI{row_idx}"] = ""
        
        # 设置数据样式
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for _, col in HUADING_FIELDS:
            cell = ws[f"{col}{row_idx}"]
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # 保存文件
    output_dir = "/tmp"
    os.makedirs(output_dir, exist_ok=True)
    output_path = f"{output_dir}/华鼎标准出库单_任丘北辛庄店_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    wb.save(output_path)
    
    print(f"✅ 华鼎标准出库单已生成: {output_path}")
    print(f"📋 订单编号: {order_header['order_no']}")
    print(f"🏪 客户名称: {order_header['customer_name']}")
    print(f"📦 商品种数: {len(order_items)}")
    print(f"📊 总数量: {sum(item['quantity'] for item in order_items)}")
    print()
    print("⚠️ 待确认字段（黄色高亮）:")
    print("  - B列: 门店编号")
    print("  - M列: 单位类型")
    print("  - P列: 配送方式")
    print("  - AD列: 业务类型")
    
    return output_path

if __name__ == "__main__":
    output_path = main()