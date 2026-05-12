# -*- coding: utf-8 -*-
"""
Excel 读取模块

负责读取华鼎系统导出的标准Excel模板数据：
- 出库单导出明细
- 仓储透视表
- 订单明细透视
- 客户信息导出
- 商品信息导出

参考PRD文档：AI账单助手 v1.1
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelReadError(Exception):
    """Excel读取异常"""
    pass


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, file_path: Union[str, Path]):
        """
        初始化Excel读取器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise ExcelReadError(f"文件不存在: {self.file_path}")
        
        self.workbook: Optional[openpyxl.Workbook] = None
        self._load_workbook()
    
    def _load_workbook(self) -> None:
        """加载工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=True,  # 读取公式计算后的值
                read_only=True
            )
            logger.info(f"成功加载Excel文件: {self.file_path}")
        except Exception as e:
            raise ExcelReadError(f"加载Excel文件失败: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """获取所有Sheet名称"""
        if self.workbook is None:
            return []
        return self.workbook.sheetnames
    
    def _get_sheet(self, sheet_name: str) -> Worksheet:
        """获取指定Sheet"""
        if self.workbook is None:
            raise ExcelReadError("工作簿未加载")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ExcelReadError(f"Sheet不存在: {sheet_name}")
        
        return self.workbook[sheet_name]
    
    @staticmethod
    def _safe_get_cell_value(cell: Cell) -> Any:
        """安全获取单元格值"""
        if cell is None:
            return None
        
        value = cell.value
        if value is None:
            return None
        
        # 处理日期类型
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        
        return value
    
    @staticmethod
    def _parse_headers(ws: Worksheet, row: int = 1) -> Dict[int, str]:
        """
        解析表头
        
        Args:
            ws: Worksheet对象
            row: 表头行号，默认第1行
            
        Returns:
            {列索引: 列名} 映射
        """
        headers = {}
        for cell in ws[row]:
            if cell.value is not None:
                headers[cell.column] = str(cell.value).strip()
        return headers
    
    @staticmethod
    def _read_data_rows(
        ws: Worksheet, 
        headers: Dict[int, str],
        start_row: int = 2,
        max_rows: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        读取数据行
        
        Args:
            ws: Worksheet对象
            headers: 表头映射
            start_row: 开始行号
            max_rows: 最大行数限制
            
        Returns:
            数据行列表
        """
        rows = []
        end_row = ws.max_row
        
        if max_rows is not None:
            end_row = min(start_row + max_rows - 1, end_row)
        
        for row_idx in range(start_row, end_row + 1):
            row_data = {}
            for col_idx, header_name in headers.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[header_name] = ExcelReader._safe_get_cell_value(cell)
            
            # 检查是否为空行
            if any(v is not None for v in row_data.values()):
                rows.append(row_data)
        
        return rows
    
    def read_sheet(
        self, 
        sheet_name: str,
        header_row: int = 1,
        data_start_row: int = 2,
        max_rows: Optional[int] = None
    ) -> Tuple[Dict[int, str], List[Dict[str, Any]]]:
        """
        读取指定Sheet的数据
        
        Args:
            sheet_name: Sheet名称
            header_row: 表头行号
            data_start_row: 数据开始行号
            max_rows: 最大行数限制
            
        Returns:
            (表头映射, 数据列表)
        """
        ws = self._get_sheet(sheet_name)
        headers = self._parse_headers(ws, header_row)
        data = self._read_data_rows(ws, headers, data_start_row, max_rows)
        return headers, data
    
    def close(self) -> None:
        """关闭工作簿"""
        if self.workbook is not None:
            self.workbook.close()
            self.workbook = None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
    
    def __del__(self):
        self.close()


class OutboundReader(ExcelReader):
    """
    出库单导出明细读取器
    
    原始数据Sheet - 华鼎系统直接导出
    
    字段定义（参考PRD 6.2）：
    | 字段名 | 类型 | 说明 | 示例 |
    | 创建日期 | date | 出库日期 | 2026-02-01 |
    | 出库仓库 | string | 始发仓库 | 成都仓 |
    | 订单类型 | string | 共配/调拨/代发 | 共配 |
    | 出库单号 | string | 唯一标识 | OCK2026020110016 |
    | 门店名称 | string | 匹配定价规则 | 宫小燕（成都星河店） |
    | 订单备注 | string | 特殊说明 | - |
    | 商品编码 | string | 关联转换比 | GG250706000013 |
    | 实际出库数量 | number | 原始数量 | 50 |
    | 计量方式 | string | 大单位/小单位 | 大单位 |
    | 存储方式 | string | 常温/冷冻/冷藏 | 冷冻 |
    | 业务日期 | date | 确定账期 | 2026-02-01 |
    | 标准件数量 | number | AI计算 | 50.47 |
    """
    
    SHEET_NAME = "出库单导出明细"
    
    # 标准表头映射（支持多种命名方式）
    HEADER_ALIASES = {
        "创建日期": ["创建日期", "出库日期", "日期", "date"],
        "出库仓库": ["出库仓库", "仓库", "始发仓库", "warehouse"],
        "订单类型": ["订单类型", "类型", "出库类型", "order_type"],
        "出库单号": ["出库单号", "单号", "订单号", "order_no"],
        "门店名称": ["门店名称", "门店", "目的门店", "store_name"],
        "订单备注": ["订单备注", "备注", "note"],
        "商品编码": ["商品编码", "SKU", "sku", "product_code"],
        "实际出库数量": ["实际出库数量", "数量", "qty", "quantity"],
        "计量方式": ["计量方式", "单位", "unit"],
        "存储方式": ["存储方式", "温区", "temperature_zone"],
        "业务日期": ["业务日期", "账期", "billing_date"],
        "标准件数量": ["标准件数量", "标准件", "std_units"],
    }
    
    def __init__(self, file_path: Union[str, Path]):
        super().__init__(file_path)
        self._headers: Optional[Dict[int, str]] = None
        self._data: Optional[List[Dict[str, Any]]] = None
    
    def read(self) -> List[Dict[str, Any]]:
        """
        读取出库单导出明细
        
        Returns:
            出库单列表
            
        Raises:
            ExcelReadError: 读取失败时抛出
        """
        try:
            if self.SHEET_NAME not in self.get_sheet_names():
                raise ExcelReadError(f"Sheet不存在: {self.SHEET_NAME}")
            
            self._headers, self._data = self.read_sheet(self.SHEET_NAME)
            logger.info(f"读取出库单导出明细: {len(self._data)}条记录")
            return self._data
            
        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(f"读取出库单导出明细失败: {e}")
    
    def get_outbound_by_order_no(self, order_no: str) -> List[Dict[str, Any]]:
        """根据出库单号获取出库明细"""
        if self._data is None:
            self.read()
        
        return [row for row in self._data if row.get("出库单号") == order_no]
    
    def get_outbound_by_store(self, store_name: str) -> List[Dict[str, Any]]:
        """根据门店名称获取出库明细"""
        if self._data is None:
            self.read()
        
        return [row for row in self._data if store_name in str(row.get("门店名称", ""))]
    
    def get_outbound_by_date(self, date: str) -> List[Dict[str, Any]]:
        """根据业务日期获取出库明细"""
        if self._data is None:
            self.read()
        
        return [row for row in self._data if row.get("业务日期") == date]
    
    def get_outbound_by_period(
        self, 
        start_date: str, 
        end_date: str
    ) -> List[Dict[str, Any]]:
        """根据日期范围获取出库明细"""
        if self._data is None:
            self.read()
        
        return [
            row for row in self._data 
            if start_date <= row.get("业务日期", "") <= end_date
        ]
    
    def aggregate_by_order(self) -> Dict[str, Dict[str, Any]]:
        """按出库单号聚合数据"""
        if self._data is None:
            self.read()
        
        aggregated = {}
        for row in self._data:
            order_no = row.get("出库单号")
            if order_no:
                if order_no not in aggregated:
                    aggregated[order_no] = {
                        "出库单号": order_no,
                        "门店名称": row.get("门店名称"),
                        "出库仓库": row.get("出库仓库"),
                        "订单类型": row.get("订单类型"),
                        "业务日期": row.get("业务日期"),
                        "明细": [],
                    }
                aggregated[order_no]["明细"].append(row)
        
        return aggregated
    
    def to_structured(self) -> Dict[str, Any]:
        """
        转换为结构化数据（用于后续计算）
        
        Returns:
            结构化出库数据
        """
        if self._data is None:
            self.read()
        
        return {
            "source": "outbound_details",
            "record_count": len(self._data),
            "records": self._data,
        }


class StorageReader(ExcelReader):
    """
    仓储透视表读取器
    
    原始数据Sheet - 华鼎系统直接导出
    
    字段定义（参考PRD 6.3）：
    | 字段名 | 类型 | 说明 | 示例 |
    | 业务日期 | date | 仓储日期 | 2026-02-01 |
    | 存储费-托盘 | number | 当日存储费 | 370.40 |
    | 统配出库费-托盘 | number | 当日统配出库费 | 1065.69 |
    | 总计 | number | 当日费用合计 | 1499.29 |
    """
    
    SHEET_NAME = "仓储透视表"
    
    def __init__(self, file_path: Union[str, Path]):
        super().__init__(file_path)
        self._data: Optional[List[Dict[str, Any]]] = None
    
    def read(self) -> List[Dict[str, Any]]:
        """
        读取仓储透视表
        
        Returns:
            仓储数据列表
            
        Raises:
            ExcelReadError: 读取失败时抛出
        """
        try:
            if self.SHEET_NAME not in self.get_sheet_names():
                raise ExcelReadError(f"Sheet不存在: {self.SHEET_NAME}")
            
            _, self._data = self.read_sheet(self.SHEET_NAME)
            
            # 过滤汇总行和空白行
            self._data = [
                row for row in self._data 
                if row.get("业务日期") and "合计" not in str(row.get("业务日期", ""))
            ]
            
            logger.info(f"读取仓储透视表: {len(self._data)}天数据")
            return self._data
            
        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(f"读取仓储透视表失败: {e}")
    
    def get_daily_storage(self, date: str) -> Optional[Dict[str, Any]]:
        """获取指定日期的仓储数据"""
        if self._data is None:
            self.read()
        
        for row in self._data:
            if row.get("业务日期") == date:
                return row
        return None
    
    def calculate_storage_fee(self) -> float:
        """计算存储费合计"""
        if self._data is None:
            self.read()
        
        total = 0.0
        for row in self._data:
            storage_fee = row.get("存储费-托盘") or 0
            if isinstance(storage_fee, str):
                storage_fee = float(storage_fee.replace(",", "")) if storage_fee else 0
            total += float(storage_fee)
        
        return round(total, 2)
    
    def calculate_outbound_fee(self) -> float:
        """计算统配出库费合计"""
        if self._data is None:
            self.read()
        
        total = 0.0
        for row in self._data:
            outbound_fee = row.get("统配出库费-托盘") or 0
            if isinstance(outbound_fee, str):
                outbound_fee = float(outbound_fee.replace(",", "")) if outbound_fee else 0
            total += float(outbound_fee)
        
        return round(total, 2)
    
    def get_period_summary(self) -> Dict[str, float]:
        """获取账期汇总"""
        if self._data is None:
            self.read()
        
        return {
            "storage_fee": self.calculate_storage_fee(),
            "outbound_fee": self.calculate_outbound_fee(),
            "total": round(self.calculate_storage_fee() + self.calculate_outbound_fee(), 2),
            "days": len(self._data),
        }
    
    def to_structured(self) -> Dict[str, Any]:
        """转换为结构化数据"""
        if self._data is None:
            self.read()
        
        return {
            "source": "storage_pivot",
            "record_count": len(self._data),
            "summary": self.get_period_summary(),
            "records": self._data,
        }


class OrderPivotReader(ExcelReader):
    """
    订单明细透视读取器
    
    计算生成Sheet - AI根据出库单+转换比+计费系数计算
    
    字段定义（参考PRD 6.4）：
    | 字段名 | 类型 | 说明 | 示例 | 计算来源 |
    | 出库单号 | string | 关联出库单 | OCK2026020110016 | 出库单导出明细 |
    | 门店名称 | string | 目的门店 | 宫小燕（成都星河店） | 出库单导出明细 |
    | 标准件数量 | number | 计算后标准件 | 50.47 | 实际数量×转换比×系数 |
    | 订单类型 | string | 发运方式 | 共配 | 出库单导出明细 |
    
    计算公式：
    标准件数量 = 实际出库数量 × 转换比（大单位→标准件）× 计费系数
    """
    
    SHEET_NAME = "订单明细透视"
    
    def __init__(self, file_path: Union[str, Path]):
        super().__init__(file_path)
        self._data: Optional[List[Dict[str, Any]]] = None
    
    def read(self) -> List[Dict[str, Any]]:
        """
        读取订单明细透视
        
        Returns:
            订单透视数据列表
            
        Raises:
            ExcelReadError: 读取失败时抛出
        """
        try:
            if self.SHEET_NAME not in self.get_sheet_names():
                raise ExcelReadError(f"Sheet不存在: {self.SHEET_NAME}")
            
            _, self._data = self.read_sheet(self.SHEET_NAME)
            logger.info(f"读取订单明细透视: {len(self._data)}条记录")
            return self._data
            
        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(f"读取订单明细透视失败: {e}")
    
    def get_order_by_no(self, order_no: str) -> Optional[Dict[str, Any]]:
        """根据出库单号获取订单数据"""
        if self._data is None:
            self.read()
        
        for row in self._data:
            if row.get("出库单号") == order_no:
                return row
        return None
    
    def calculate_transport_fee(
        self,
        order_no: str,
        unit_price: float,
        min_fee: float
    ) -> float:
        """
        计算运输费
        
        公式: MAX(标准件数量 × 单价, 最低收费)
        
        Args:
            order_no: 出库单号
            unit_price: 单价（元/件）
            min_fee: 最低收费
            
        Returns:
            运输费
        """
        order = self.get_order_by_no(order_no)
        if order is None:
            raise ExcelReadError(f"订单不存在: {order_no}")
        
        std_units = order.get("标准件数量") or 0
        if isinstance(std_units, str):
            std_units = float(std_units.replace(",", "")) if std_units else 0
        
        calculated_fee = float(std_units) * unit_price
        return round(max(calculated_fee, min_fee), 2)
    
    def to_structured(self) -> Dict[str, Any]:
        """转换为结构化数据"""
        if self._data is None:
            self.read()
        
        return {
            "source": "order_pivot",
            "record_count": len(self._data),
            "records": self._data,
        }


class CustomerInfoReader(ExcelReader):
    """
    客户信息导出读取器
    
    计算生成Sheet - AI根据合同规则+门店匹配生成
    
    字段定义（参考PRD 6.5）：
    | 字段名 | 类型 | 说明 | 示例 | 计算来源 |
    | 门店名称 | string | 完整门店名 | 宫小燕（成都星河店） | 来源：门店档案 |
    | 最低收费 | number | 最低收费金额 | 40 | 计算：合同规则→仓库→区域报价 |
    | 单价 | number | 单价（元/件） | 3.5 | 计算：合同规则→仓库→区域报价 |
    | 调拨单价 | number | 调拨单价 | 8 | 计算：合同规则→调拨费率 |
    | 调拨段运费 | number | 调拨段运费计算值 | - | 计算：标准件×调拨单价 |
    """
    
    SHEET_NAME = "客户信息导出"
    
    def __init__(self, file_path: Union[str, Path]):
        super().__init__(file_path)
        self._data: Optional[List[Dict[str, Any]]] = None
    
    def read(self) -> List[Dict[str, Any]]:
        """
        读取客户信息导出
        
        Returns:
            客户信息列表
            
        Raises:
            ExcelReadError: 读取失败时抛出
        """
        try:
            if self.SHEET_NAME not in self.get_sheet_names():
                raise ExcelReadError(f"Sheet不存在: {self.SHEET_NAME}")
            
            _, self._data = self.read_sheet(self.SHEET_NAME)
            logger.info(f"读取客户信息导出: {len(self._data)}条记录")
            return self._data
            
        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(f"读取客户信息导出失败: {e}")
    
    def get_store_pricing(self, store_name: str) -> Optional[Dict[str, Any]]:
        """获取门店定价信息"""
        if self._data is None:
            self.read()
        
        for row in self._data:
            if row.get("门店名称") == store_name:
                return {
                    "门店名称": row.get("门店名称"),
                    "最低收费": self._parse_number(row.get("最低收费")),
                    "单价": self._parse_number(row.get("单价")),
                    "调拨单价": self._parse_number(row.get("调拨单价")),
                }
        return None
    
    @staticmethod
    def _parse_number(value: Any) -> float:
        """解析数字"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            return float(value.replace(",", "")) if value else 0.0
        return 0.0
    
    def build_pricing_lookup(self) -> Dict[str, Dict[str, float]]:
        """
        构建定价查询表
        
        Returns:
            {门店名称: {最低收费, 单价, 调拨单价}}
        """
        if self._data is None:
            self.read()
        
        lookup = {}
        for row in self._data:
            store_name = row.get("门店名称")
            if store_name:
                lookup[store_name] = {
                    "最低收费": self._parse_number(row.get("最低收费")),
                    "单价": self._parse_number(row.get("单价")),
                    "调拨单价": self._parse_number(row.get("调拨单价")),
                }
        
        return lookup
    
    def to_structured(self) -> Dict[str, Any]:
        """转换为结构化数据"""
        if self._data is None:
            self.read()
        
        return {
            "source": "customer_info",
            "record_count": len(self._data),
            "pricing_lookup": self.build_pricing_lookup(),
            "records": self._data,
        }


class ProductInfoReader(ExcelReader):
    """
    商品信息导出读取器
    
    计算生成Sheet - AI根据合同标准件定义+商品档案计算
    
    字段定义（参考PRD 6.6）：
    | 字段名 | 类型 | 说明 | 示例 | 计算来源 |
    | 商品编码 | string | 商品唯一码 | GG250706000013 | 来源：商品档案 |
    | 商品名称 | string | 商品名称 | 500ml糖水瓶72个/箱 | 来源：商品档案 |
    | 单件重量 | number | 单件重量(kg) | 15kg | 来源：商品档案 |
    | 单件体积 | number | 单件体积(m³) | 0.05m³ | 来源：商品档案 |
    | 计费系数 | number | 标准件系数 | 1.0 | 计算：合同标准件定义 |
    
    计算公式（合同标准件定义）：
    系数 = 1.0（基础）
        + 0.2（若 15kg < 重量 ≤ 25kg 或 0.05m³ < 体积 ≤ 0.08m³）
        + 0.2（若 25kg < 重量 ≤ 50kg 或 0.08m³ < 体积 ≤ 0.1m³）
        + 0.2×N（每再增加5kg或0.05m³）
    """
    
    SHEET_NAME = "商品信息导出"
    
    def __init__(self, file_path: Union[str, Path]):
        super().__init__(file_path)
        self._data: Optional[List[Dict[str, Any]]] = None
    
    def read(self) -> List[Dict[str, Any]]:
        """
        读取商品信息导出
        
        Returns:
            商品信息列表
            
        Raises:
            ExcelReadError: 读取失败时抛出
        """
        try:
            if self.SHEET_NAME not in self.get_sheet_names():
                raise ExcelReadError(f"Sheet不存在: {self.SHEET_NAME}")
            
            _, self._data = self.read_sheet(self.SHEET_NAME)
            logger.info(f"读取商品信息导出: {len(self._data)}条记录")
            return self._data
            
        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(f"读取商品信息导出失败: {e}")
    
    def get_product_coefficient(self, product_code: str) -> float:
        """获取商品计费系数"""
        if self._data is None:
            self.read()
        
        for row in self._data:
            if row.get("商品编码") == product_code:
                return self._parse_number(row.get("计费系数"))
        return 1.0  # 默认系数
    
    @staticmethod
    def _parse_number(value: Any) -> float:
        """解析数字"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            return float(value.replace(",", "")) if value else 0.0
        return 0.0
    
    @staticmethod
    def calculate_coefficient(weight: float, volume: float) -> float:
        """
        计算计费系数（根据合同标准件定义）
        
        Args:
            weight: 单件重量(kg)
            volume: 单件体积(m³)
            
        Returns:
            计费系数
        """
        coefficient = 1.0
        
        # 重量超标检测
        if weight > 15:
            if weight <= 25:
                coefficient += 0.2
            elif weight <= 50:
                coefficient += 0.4
            else:
                # 每+5kg +0.2
                extra_weight = weight - 50
                extra_tiers = extra_weight // 5
                if extra_weight % 5 > 0:
                    extra_tiers += 1
                coefficient += 0.4 + extra_tiers * 0.2
        
        # 体积超标检测
        if volume > 0.05:
            if volume <= 0.08:
                coefficient += 0.2
            elif volume <= 0.1:
                coefficient += 0.4
            else:
                # 每+0.05m³ +0.2
                extra_volume = volume - 0.1
                extra_tiers = extra_volume // 0.05
                if extra_volume % 0.05 > 0:
                    extra_tiers += 1
                coefficient += 0.4 + extra_tiers * 0.2
        
        return round(coefficient, 2)
    
    def build_coefficient_lookup(self) -> Dict[str, float]:
        """
        构建系数查询表
        
        Returns:
            {商品编码: 计费系数}
        """
        if self._data is None:
            self.read()
        
        lookup = {}
        for row in self._data:
            product_code = row.get("商品编码")
            if product_code:
                lookup[product_code] = self._parse_number(row.get("计费系数"))
        
        return lookup
    
    def to_structured(self) -> Dict[str, Any]:
        """转换为结构化数据"""
        if self._data is None:
            self.read()
        
        return {
            "source": "product_info",
            "record_count": len(self._data),
            "coefficient_lookup": self.build_coefficient_lookup(),
            "records": self._data,
        }


# ============================================================
# 便捷函数
# ============================================================

def read_outbound_details(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    """
    读取出库单导出明细
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        出库单列表
    """
    with OutboundReader(file_path) as reader:
        return reader.read()


def read_storage_pivot(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    """
    读取仓储透视表
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        仓储数据列表
    """
    with StorageReader(file_path) as reader:
        return reader.read()


def read_order_pivot(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    """
    读取订单明细透视
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        订单透视数据列表
    """
    with OrderPivotReader(file_path) as reader:
        return reader.read()


def read_customer_info(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    """
    读取客户信息导出
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        客户信息列表
    """
    with CustomerInfoReader(file_path) as reader:
        return reader.read()


def read_product_info(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    """
    读取商品信息导出
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        商品信息列表
    """
    with ProductInfoReader(file_path) as reader:
        return reader.read()


def read_all_excel_data(file_path: Union[str, Path]) -> Dict[str, Any]:
    """
    读取Excel所有数据（出库单 + 仓储）
    
    适用于MVP阶段：客服从华鼎系统导出包含出库明细和仓储透视表的单个Excel文件
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        所有数据的结构化汇总
    """
    result = {
        "source_file": str(file_path),
        "read_time": datetime.now().isoformat(),
        "data": {}
    }
    
    with OutboundReader(file_path) as outbound_reader:
        result["data"]["outbound"] = outbound_reader.to_structured()
    
    with StorageReader(file_path) as storage_reader:
        result["data"]["storage"] = storage_reader.to_structured()
    
    return result


if __name__ == "__main__":
    # 测试代码
    import sys
    
    if len(sys.argv) > 1:
        test_file = sys.argv[1]
        print(f"测试读取文件: {test_file}")
        
        try:
            # 测试读取出库单
            outbound_data = read_outbound_details(test_file)
            print(f"出库单记录数: {len(outbound_data)}")
            if outbound_data:
                print(f"第一条记录: {outbound_data[0]}")
            
        except ExcelReadError as e:
            print(f"读取错误: {e}")
    else:
        print("用法: python reader.py <excel_file_path>")
