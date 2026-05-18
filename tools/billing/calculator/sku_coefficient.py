# billing_tools/calculator/sku_coefficient.py
# SKU计费系数计算器 v2.3

"""
SKU计费系数计算

根据商品信息和计费规则，计算每个SKU的计费系数

功能：
1. 读取商品信息Excel
2. 从规则库获取：标准件定义、固定SKU列表、超规格规则（从合同提取）
3. 对每个SKU计算计费系数：
   - 固定SKU → 系数 = 1.0
   - 非固定SKU → 按合同规定的超规格规则计算
4. 在原Excel加一列"计费系数"，计算后覆盖

超规格规则从合同提取（rules.oversized），包含：
- 基准重量、基准体积
- 各等级阈值（tier1/tier2/tier3 的重量和体积边界）
- 各等级系数（tier1系数、tier2系数）
- tier3使用公式 max(重量/基准重量, 体积/基准体积)
"""

import sys
import os
import json

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from knowledge.rule_store import load_rule

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None


def calc_sku_coefficients(
    customer_id: str,
    product_excel_path: str,
    output_excel_path: str = None
) -> dict:
    """
    计算SKU计费系数
    
    Args:
        customer_id: 客户ID
        product_excel_path: 商品信息Excel文件路径
        output_excel_path: 输出Excel路径（可选，默认覆盖原文件）
        
    Returns:
        {
            "success": bool,
            "product_count": int,       # 商品总数
            "fixed_sku_count": int,       # 固定SKU数量（系数=1.0且来自合同）
            "calculated_count": int,      # 计算SKU数量
            "coefficients": dict,          # {商品编码: 计费系数}
            "message": str
        }
    """
    if not openpyxl:
        return {
            "success": False,
            "message": "需要安装 openpyxl: pip install openpyxl"
        }
    
    if not os.path.exists(product_excel_path):
        return {
            "success": False,
            "message": f"文件不存在: {product_excel_path}"
        }
    
    # 1. 加载规则库
    rule_data = load_rule(customer_id)
    if not rule_data:
        return {
            "success": False,
            "message": f"未找到客户规则: {customer_id}"
        }
    
    rules = rule_data.get('rules', {})
    product_coefficients = rules.get('product_coefficients', {})
    standard_unit_def = product_coefficients.get('standard_unit_definition', {})
    fixed_skus = product_coefficients.get('fixed_skus', {})
    # 超规格规则：优先从 rules.oversized 读取（v2.1+），兼容 product_coefficients.oversized_rules（旧版）
    oversized_rules = rules.get('oversized') or product_coefficients.get('oversized_rules', {})
    
    # 标准件定义默认值
    max_weight = standard_unit_def.get('max_weight_kg', 15)
    max_volume = standard_unit_def.get('max_volume_m3', 0.05)
    
    # 2. 读取商品信息Excel
    wb = openpyxl.load_workbook(product_excel_path)
    ws = wb.active
    
    # 查找表头行
    header_row = 1
    headers = {}
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row, col).value or '').strip()
            if '商品编码' in cell_value:
                headers['商品编码'] = col
            elif '商品名称' in cell_value:
                headers['商品名称'] = col
            elif '商品标准单位体积' in cell_value:
                headers['体积'] = col
            elif '商品标准单位重量' in cell_value:
                headers['重量'] = col
            elif '计费系数' in cell_value:
                headers['计费系数'] = col
    
    if '商品编码' not in headers:
        return {
            "success": False,
            "message": "未找到'商品编码'列"
        }
    
    # 3. 确定计费系数列
    coef_col = headers.get('计费系数')
    if coef_col is None:
        # 在最后一列后面添加
        coef_col = ws.max_column + 1
        ws.cell(1, coef_col).value = '计费系数'
        ws.cell(1, coef_col).font = Font(bold=True)
    
    # 4. 遍历商品，计算计费系数
    coefficients = {}
    fixed_count = 0
    calculated_count = 0
    skipped_count = 0
    
    for row in range(2, ws.max_row + 1):
        sku_code = str(ws.cell(row, headers['商品编码']).value or '').strip()
        if not sku_code or sku_code == 'None':
            continue
        
        # 获取重量和体积
        weight_kg = None
        volume_m3 = None
        
        if '重量' in headers:
            weight_val = ws.cell(row, headers['重量']).value
            if weight_val is not None:
                try:
                    weight_kg = float(weight_val)
                except (ValueError, TypeError):
                    pass
        
        if '体积' in headers:
            # Excel中的体积单位可能是 cm³，需要转换为 m³
            volume_val = ws.cell(row, headers['体积']).value
            if volume_val is not None:
                try:
                    volume_cm3 = float(volume_val)
                    volume_m3 = volume_cm3 / 1000000  # cm³ → m³
                except (ValueError, TypeError):
                    pass
        
        # 计算计费系数
        coef = calculate_coefficient(
            sku_code,
            weight_kg,
            volume_m3,
            fixed_skus,
            max_weight,
            max_volume,
            oversized_rules
        )
        
        coefficients[sku_code] = coef
        
        # 写入Excel
        ws.cell(row, coef_col).value = coef
        
        if coef == 1.0 and sku_code in fixed_skus:
            fixed_count += 1
        else:
            calculated_count += 1
    
    # 5. 保存Excel
    if output_excel_path:
        save_path = output_excel_path
    else:
        save_path = product_excel_path
    
    wb.save(save_path)
    
    # 6. 保存商品信息表到规则库（供费用计算使用）
    _save_product_table(customer_id, coefficients, product_excel_path)
    
    total_count = fixed_count + calculated_count
    
    return {
        "success": True,
        "product_count": total_count,
        "fixed_sku_count": fixed_count,
        "calculated_count": calculated_count,
        "coefficients": coefficients,
        "message": f"计算完成：共{total_count}个SKU，固定{fixed_count}个，计算{calculated_count}个"
    }


def calculate_coefficient(
    sku_code: str,
    weight_kg: float,
    volume_m3: float,
    fixed_skus: dict,
    max_weight: float = 15,
    max_volume: float = 0.05,
    oversized_rules: dict = None
) -> float:
    """
    计算单个SKU的计费系数
    
    Args:
        sku_code: 商品编码
        weight_kg: 商品重量（kg）
        volume_m3: 商品体积（m³）
        fixed_skus: 固定SKU字典 {sku_code: {name, spec, volume_m3, weight_kg}}
        max_weight: 标准件最大重量（kg）
        max_volume: 标准件最大体积（m³）
        oversized_rules: 超规格规则
        
    Returns:
        计费系数（float）
    """
    # 1. 检查是否是固定SKU
    if sku_code in fixed_skus:
        return 1.0
    
    # 2. 如果没有重量和体积信息，默认1.0
    if weight_kg is None and volume_m3 is None:
        return 1.0
    
    # 3. 如果没有重量或体积，用已有的那个计算
    if weight_kg is None:
        weight_kg = max_weight  # 如果没有重量，按标准件重量计算
    if volume_m3 is None:
        volume_m3 = max_volume  # 如果没有体积，按标准件体积计算
    
    # 4. 计算超规格系数
    coef = compute_oversized_coefficient(
        weight_kg, volume_m3, max_weight, max_volume, oversized_rules
    )
    
    return coef


def compute_oversized_coefficient(
    weight_kg: float,
    volume_m3: float,
    max_weight: float = 15,
    max_volume: float = 0.05,
    oversized_rules: dict = None
) -> float:
    """
    计算超规格系数
    
    超规格规则从规则库提取（rules.oversized），结构如下：
    {
        "基准重量_kg": 15,
        "基准体积_m3": 0.05,
        "threshold_tier1": {"weight_kg": {"min": 15, "max": 25}, "volume_m3": {"min": 0.05, "max": 0.08}},
        "threshold_tier2": {"weight_kg": {"min": 25, "max": 30}, "volume_m3": {"min": 0.08, "max": 0.1}},
        "threshold_tier3": {"weight_kg": {"min": 30}, "volume_m3": {"min": 0.1}},
        "coef_tier1": 1.2,
        "coef_tier2": 2.0
    }
    
    判断逻辑：任一维度达到更高等级，即用更高等级的系数
    """
    # 如果没有超规格规则，使用默认值（兼容旧规则库）
    if not oversized_rules:
        weight_ratio = weight_kg / max_weight if max_weight > 0 else 0
        volume_ratio = volume_m3 / max_volume if max_volume > 0 else 0
        
        # tier3: 重量>30kg 或 体积>0.1m³
        if volume_m3 > 0.1 or weight_kg > 30:
            return max(weight_ratio, volume_ratio)
        # tier2: 25kg<重量≤30kg 或 0.08m³<体积≤0.1m³
        elif volume_m3 > 0.08 or weight_kg > 25:
            return 2.0
        # tier1: 15kg<重量≤25kg 或 0.05m³<体积≤0.08m³
        elif volume_m3 > 0.05 or weight_kg > 15:
            return 1.2
        else:
            return 1.0
    
    # 从规则库读取阈值
    base_weight = oversized_rules.get('基准重量_kg', max_weight)
    base_volume = oversized_rules.get('基准体积_m3', max_volume)
    
    tier1 = oversized_rules.get('threshold_tier1', {})
    tier2 = oversized_rules.get('threshold_tier2', {})
    tier3 = oversized_rules.get('threshold_tier3', {})
    
    coef_tier1 = oversized_rules.get('coef_tier1', 1.2)
    coef_tier2 = oversized_rules.get('coef_tier2', 2.0)
    
    # 提取各等级阈值
    t1_weight_min = tier1.get('weight_kg', {}).get('min', 15)
    t1_weight_max = tier1.get('weight_kg', {}).get('max', 25)
    t1_volume_min = tier1.get('volume_m3', {}).get('min', 0.05)
    t1_volume_max = tier1.get('volume_m3', {}).get('max', 0.08)
    
    t2_weight_min = tier2.get('weight_kg', {}).get('min', 25)
    t2_weight_max = tier2.get('weight_kg', {}).get('max', 50)
    t2_volume_min = tier2.get('volume_m3', {}).get('min', 0.08)
    t2_volume_max = tier2.get('volume_m3', {}).get('max', 0.1)
    
    t3_weight_min = tier3.get('weight_kg', {}).get('min', 50)
    t3_volume_min = tier3.get('volume_m3', {}).get('min', 0.1)
    
    # 判断超规格等级（任一维度达到更高等级，即用更高等级）
    # tier3: 重量>t3_min 或 体积>t3_min
    if volume_m3 > t3_volume_min or weight_kg > t3_weight_min:
        weight_ratio = weight_kg / base_weight if base_weight > 0 else 0
        volume_ratio = volume_m3 / base_volume if base_volume > 0 else 0
        return max(weight_ratio, volume_ratio)
    # tier2: (重量>t2_min 或 体积>t2_min) 且 未达到 tier3
    elif (volume_m3 > t2_volume_min or weight_kg > t2_weight_min):
        return coef_tier2
    # tier1: (重量>t1_min 或 体积>t1_min) 且 未达到 tier2/tier3
    elif (volume_m3 > t1_volume_min or weight_kg > t1_weight_min):
        return coef_tier1
    else:
        return 1.0


def batch_calc_sku_coefficients(
    customer_id: str,
    product_excel_path: str,
    output_excel_path: str = None,
    skip_existing: bool = True
) -> dict:
    """
    批量计算SKU计费系数（跳过已有系数的行）
    
    Args:
        customer_id: 客户ID
        product_excel_path: 商品信息Excel文件路径
        output_excel_path: 输出Excel路径
        skip_existing: 是否跳过已有计费系数的行
        
    Returns:
        同 calc_sku_coefficients
    """
    if not openpyxl:
        return {"success": False, "message": "需要安装 openpyxl"}
    
    if not os.path.exists(product_excel_path):
        return {"success": False, "message": f"文件不存在: {product_excel_path}"}
    
    # 加载规则
    rule_data = load_rule(customer_id)
    if not rule_data:
        return {"success": False, "message": f"未找到客户规则: {customer_id}"}
    
    rules = rule_data.get('rules', {})
    product_coefficients = rules.get('product_coefficients', {})
    fixed_skus = product_coefficients.get('fixed_skus', {})
    standard_unit_def = product_coefficients.get('standard_unit_definition', {})
    
    max_weight = standard_unit_def.get('max_weight_kg', 15)
    max_volume = standard_unit_def.get('max_volume_m3', 0.05)
    
    # 读取Excel
    wb = openpyxl.load_workbook(product_excel_path)
    ws = wb.active
    
    # 查找表头
    headers = {}
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row, col).value or '').strip()
            if '商品编码' in cell_value:
                headers['商品编码'] = col
            elif '商品名称' in cell_value:
                headers['商品名称'] = col
            elif '商品标准单位体积' in cell_value:
                headers['体积'] = col
            elif '商品标准单位重量' in cell_value:
                headers['重量'] = col
            elif '计费系数' in cell_value:
                headers['计费系数'] = col
    
    if '商品编码' not in headers:
        return {"success": False, "message": "未找到'商品编码'列"}
    
    # 确定计费系数列
    coef_col = headers.get('计费系数')
    if coef_col is None:
        coef_col = ws.max_column + 1
        ws.cell(1, coef_col).value = '计费系数'
        ws.cell(1, coef_col).font = Font(bold=True)
    
    # 遍历计算
    coefficients = {}
    fixed_count = 0
    calculated_count = 0
    
    for row in range(2, ws.max_row + 1):
        sku_code = str(ws.cell(row, headers['商品编码']).value or '').strip()
        if not sku_code or sku_code == 'None':
            continue
        
        # 如果跳过已有系数且已有值
        if skip_existing and coef_col in headers:
            existing_coef = ws.cell(row, coef_col).value
            if existing_coef is not None:
                try:
                    float(existing_coef)
                    continue  # 已有系数，跳过
                except (ValueError, TypeError):
                    pass
        
        # 获取重量体积
        weight_kg = None
        volume_m3 = None
        
        if '重量' in headers:
            weight_val = ws.cell(row, headers['重量']).value
            if weight_val is not None:
                try:
                    weight_kg = float(weight_val)
                except:
                    pass
        
        if '体积' in headers:
            volume_val = ws.cell(row, headers['体积']).value
            if volume_val is not None:
                try:
                    volume_cm3 = float(volume_val)
                    volume_m3 = volume_cm3 / 1000000
                except:
                    pass
        
        # 计算系数
        coef = calculate_coefficient(sku_code, weight_kg, volume_m3, fixed_skus, max_weight, max_volume)
        
        coefficients[sku_code] = coef
        ws.cell(row, coef_col).value = coef
        
        if sku_code in fixed_skus:
            fixed_count += 1
        else:
            calculated_count += 1
    
    # 保存
    save_path = output_excel_path or product_excel_path
    wb.save(save_path)
    
    return {
        "success": True,
        "product_count": fixed_count + calculated_count,
        "fixed_sku_count": fixed_count,
        "calculated_count": calculated_count,
        "coefficients": coefficients,
        "message": f"计算完成：共{fixed_count + calculated_count}个SKU"
    }
