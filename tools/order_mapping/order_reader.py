# -*- coding: utf-8 -*-
"""
订单读取模块 - 华鼎格式订单解析

支持华鼎系统导出的标准订单格式
"""

import os
from typing import Dict, List, Any, Optional


def read_order(file_path: str) -> Dict[str, Any]:
    """
    读取华鼎格式订单Excel文件
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        {
            "success": bool,
            "order_id": str,
            "customer_name": str,
            "store_name": str,
            "store_address": str,
            "items": [
                {
                    "row_index": int,
                    "sku_code": str,
                    "sku_name": str,
                    "quantity": float,
                    "unit": str,
                    "price": float,
                    "spec": str,
                    "notes": str
                }
            ],
            "headers": List[str],
            "row_count": int,
            "message": str
        }
    """
    try:
        import openpyxl
        
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": f"文件不存在: {file_path}",
                "items": []
            }
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 找到正确的Sheet
        sheet = None
        for sheet_name in wb.sheetnames:
            if '订货' in sheet_name or 'Sheet' in sheet_name:
                sheet = wb[sheet_name]
                break
        
        if not sheet:
            sheet = wb.active
        
        result = parse_order_sheet_v2(sheet)
        
        wb.close()
        
        return result
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "items": []
        }


def parse_order_sheet_v2(sheet) -> Dict[str, Any]:
    """解析订单Sheet - 改进版"""
    all_rows = list(sheet.iter_rows(values_only=True))
    total_rows = len(all_rows)
    
    # 初始化变量
    order_id = ""
    customer_name = ""
    store_name = ""
    store_address = ""
    contact = ""
    phone = ""
    items = []
    
    # 第一遍：查找基本信息和表头位置
    header_row_idx = -1
    data_start_row_idx = -1
    
    for row_idx, row in enumerate(all_rows):
        if not row:
            continue
            
        # 检查是否包含"商品明细"标记（注意可能有空格分隔）
        row_str = str(row).replace(' ', '')
        if "商品明细" in row_str:
            # 表头在下一行，数据从下下行开始
            header_row_idx = row_idx + 1
            data_start_row_idx = row_idx + 2
            continue
        
        # 在商品明细之前查找基本信息
        if header_row_idx < 0:
            first_cell = str(row[0]).replace(' ', '') if row[0] else ""
            
            # 解析基本信息
            if "订单编号" in first_cell and len(row) > 1 and row[1]:
                order_id = str(row[1])
            elif "公司名称" in first_cell or "门店" in first_cell:
                if len(row) > 1 and row[1]:
                    customer_name = str(row[1])
                    store_name = customer_name
            elif "收货地址" in first_cell or "地址" in first_cell:
                if len(row) > 1 and row[1]:
                    store_address = str(row[1])
                if len(row) > 6 and row[6]:
                    phone = str(row[6])
            elif "联系人" in first_cell and len(row) > 1 and row[1]:
                contact = str(row[1])
    
    # 第二遍：解析商品数据
    if data_start_row_idx > 0 and data_start_row_idx < total_rows:
        for row_idx in range(data_start_row_idx, total_rows):
            row = all_rows[row_idx]
            
            # 跳过空行
            if not row or (row[0] is None and row[1] is None and row[2] is None):
                continue
            
            # 解析商品行
            item = parse_item_row(row, row_idx)
            if item:
                items.append(item)
    
    return {
        "success": True,
        "order_id": order_id,
        "customer_name": customer_name,
        "store_name": store_name,
        "store_address": store_address,
        "contact": contact,
        "phone": phone,
        "items": items,
        "headers": ["序号", "商品编码", "商品名称", "商品规格", "", "数量", "计量单位", "", "备注"],
        "row_count": len(items),
        "message": f"读取成功，订单号:{order_id}，共{len(items)}条商品"
    }


def parse_item_row(row, row_idx: int) -> Optional[Dict]:
    """解析商品数据行"""
    try:
        # 获取各列数据
        seq = row[0]  # 序号
        sku_code = row[1] if len(row) > 1 else None  # 商品编码
        sku_name = row[2] if len(row) > 2 else None  # 商品名称
        spec = row[3] if len(row) > 3 else None  # 规格
        quantity = row[5] if len(row) > 5 else None  # 数量
        unit = row[6] if len(row) > 6 else None  # 单位
        notes = row[8] if len(row) > 8 else None  # 备注
        
        # 跳过空行
        if sku_code is None and sku_name is None:
            return None
        
        # 跳过标题行
        if sku_code == "商品编码" or sku_code == "序号":
            return None
        
        # 转换数量
        qty = 0
        if quantity is not None:
            try:
                qty = float(quantity)
            except (ValueError, TypeError):
                qty = 0
        
        return {
            "row_index": row_idx,
            "sku_code": str(sku_code).strip() if sku_code else "",
            "sku_name": str(sku_name).strip() if sku_name else "",
            "quantity": qty,
            "unit": str(unit).strip() if unit else "",
            "price": 0.0,
            "spec": str(spec).strip() if spec else "",
            "notes": str(notes).strip() if notes else ""
        }
        
    except Exception as e:
        return None


def get_order_info(file_path: str) -> Dict[str, Any]:
    """快速获取订单基本信息"""
    result = read_order(file_path)
    
    return {
        "order_id": result.get("order_id", ""),
        "customer_name": result.get("customer_name", ""),
        "store_name": result.get("store_name", ""),
        "item_count": result.get("row_count", 0),
        "file_name": os.path.basename(file_path)
    }


if __name__ == "__main__":
    # 测试
    test_file = "/Users/jinqianfei/Library/Mobile Documents/com~apple~CloudDocs/ProductManagement/AI客服赋能/资料/沧州任丘北辛庄店 肖文月30件(1).xlsx"
    
    print("=" * 60)
    print("订单读取测试")
    print("=" * 60)
    
    result = read_order(test_file)
    
    print(f"\n✅ 成功: {result['success']}")
    print(f"📋 订单号: {result.get('order_id', 'N/A')}")
    print(f"👤 客户名: {result.get('customer_name', 'N/A')}")
    print(f"🏪 门店名: {result.get('store_name', 'N/A')}")
    print(f"📍 地址: {result.get('store_address', 'N/A')[:40]}")
    print(f"📱 电话: {result.get('phone', 'N/A')}")
    print(f"📊 商品数: {result.get('row_count', 0)}")
    
    print(f"\n【商品明细】")
    for i, item in enumerate(result.get('items', [])):
        print(f"  {i+1:2d}. {item['sku_code']:15s} {item['sku_name'][:25]:25s} × {item['quantity']}{item['unit']}")
    
    print("\n" + "=" * 60)
