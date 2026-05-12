"""
华鼎模板生成器
生成华鼎31字段Excel模板
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, List, Any, Optional
import os


# 华鼎模板31个字段
HUADING_FIELDS = [
    "序号",
    "门店编号",
    "门店三方编码",
    "仓库编码",
    "加急程度",
    "商品SKU编号",
    "商品三方SPEC编号",
    "单位类型",
    "出库数量",
    "指定库存状态",
    "出库类型",
    "配送方式",
    "指定车型（专配）",
    "是否垫付",
    "付款方式",
    "快递公司",
    "单价",
    "总金额",
    "是否制定批次",
    "批次号",
    "生产日期",
    "备注",
    "生产厂家编号",
    "门店收货地址编码",
    "三方单号",
    "业务模式",
    "业务模式备注",
    "收件人",
    "收件人电话",
    "收件人地址",
    "特殊要求",
]


# 默认值配置
DEFAULT_VALUES = {
    "加急程度": 0,
    "指定库存状态": "正常",
    "出库类型": 201,
    "业务模式": "B2C",
    "是否制定批次": "否",
    "付款方式": "预付",
}


def _get_default_value(field_name: str) -> Any:
    """获取字段默认值"""
    return DEFAULT_VALUES.get(field_name, "")


def _create_header_style():
    """创建表头样式"""
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return {"font": font, "fill": fill, "alignment": alignment, "border": border}


def _create_cell_style():
    """创建单元格样式"""
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return {"alignment": alignment, "border": border}


def generate_huating_template(order_mapping_result: List[Dict[str, Any]], output_path: str) -> str:
    """
    生成华鼎31字段Excel模板

    Args:
        order_mapping_result: 订单映射结果列表，每项为一个订单的字段映射字典
        output_path: 输出文件路径

    Returns:
        输出文件的绝对路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "华鼎订单模板"

    # 设置表头
    header_style = _create_header_style()
    for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field_name)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.alignment = header_style["alignment"]
        cell.border = header_style["border"]

    # 设置列宽
    column_widths = {
        "序号": 8,
        "门店编号": 15,
        "门店三方编码": 15,
        "仓库编码": 12,
        "加急程度": 10,
        "商品SKU编号": 18,
        "商品三方SPEC编号": 18,
        "单位类型": 10,
        "出库数量": 12,
        "指定库存状态": 12,
        "出库类型": 10,
        "配送方式": 12,
        "指定车型（专配）": 15,
        "是否垫付": 10,
        "付款方式": 10,
        "快递公司": 12,
        "单价": 12,
        "总金额": 12,
        "是否制定批次": 12,
        "批次号": 18,
        "生产日期": 15,
        "备注": 25,
        "生产厂家编号": 18,
        "门店收货地址编码": 18,
        "三方单号": 20,
        "业务模式": 10,
        "业务模式备注": 20,
        "收件人": 12,
        "收件人电话": 15,
        "收件人地址": 35,
        "特殊要求": 25,
    }

    for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
        width = column_widths.get(field_name, 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 填充数据
    cell_style = _create_cell_style()

    # 提取订单数据（支持OrderMappingResult或list）
    if hasattr(order_mapping_result, 'items'):
        # OrderMappingResult对象
        order_items = order_mapping_result.items
        store_data = order_mapping_result.store or {}
        warehouse_data = order_mapping_result.warehouse or {}
    else:
        # 直接是列表
        order_items = order_mapping_result
        store_data = {}
        warehouse_data = {}

    for row_idx, item_wrapper in enumerate(order_items, start=2):
        # 获取mapped数据或原始数据
        if isinstance(item_wrapper, dict) and 'mapped' in item_wrapper:
            mapped = item_wrapper['mapped']
            original = item_wrapper['original']
            sku_code = mapped.sku_code if mapped else None
            sku_name = mapped.sku_name if mapped else original.get('product', '')
            quantity = original.get('quantity', 0)
            unit = mapped.unit_type if mapped else original.get('unit', '')
        else:
            sku_code = item_wrapper.get('sku_code', '')
            sku_name = item_wrapper.get('sku_name', item_wrapper.get('product', ''))
            quantity = item_wrapper.get('quantity', 0)
            unit = item_wrapper.get('unit_type', item_wrapper.get('unit', ''))

        for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
            # 根据字段名填充数据
            if field_name == '序号':
                value = row_idx - 1
            elif field_name == '门店编号':
                value = store_data.get('huading_store_code', '')
            elif field_name == '仓库编码':
                value = warehouse_data.get('warehouse_code', '')
            elif field_name == '商品SKU编号':
                value = sku_code or ''
            elif field_name == '单位类型':
                value = unit or ''
            elif field_name == '出库数量':
                value = quantity
            else:
                value = _get_default_value(field_name)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_style["alignment"]
            cell.border = cell_style["border"]

    # 确保输出目录存在
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # 保存文件
    wb.save(output_path)

    # 返回绝对路径
    return os.path.abspath(output_path)


def create_empty_template(output_path: str) -> str:
    """
    创建空白华鼎模板（仅表头，无数据行）

    Args:
        output_path: 输出文件路径

    Returns:
        输出文件的绝对路径
    """
    return generate_huating_template([], output_path)


# 示例用法
if __name__ == "__main__":
    # 示例订单数据
    sample_orders = [
        {
            "序号": 1,
            "门店编号": "MD001",
            "门店三方编码": "SF001",
            "仓库编码": "CK001",
            "商品SKU编号": "SKU123456",
            "商品三方SPEC编号": "SPEC001",
            "单位类型": "件",
            "出库数量": 10,
            "单价": 25.5,
            "总金额": 255.0,
            "收件人": "张三",
            "收件人电话": "13800138000",
            "收件人地址": "北京市朝阳区某某路123号",
        },
        {
            "序号": 2,
            "门店编号": "MD002",
            "门店三方编码": "SF002",
            "仓库编码": "CK001",
            "商品SKU编号": "SKU789012",
            "商品三方SPEC编号": "SPEC002",
            "单位类型": "箱",
            "出库数量": 5,
            "单价": 100.0,
            "总金额": 500.0,
            "收件人": "李四",
            "收件人电话": "13900139000",
            "收件人地址": "上海市浦东新区某某街456号",
        },
    ]

    # 生成示例文件
    output = generate_huating_template(sample_orders, "./huading_template.xlsx")
    print(f"模板已生成: {output}")
