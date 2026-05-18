"""
skill-order-to-huading-template

将客户订单Excel转换为华鼎出库单模板（31字段）
支持Excel/图片/PDF/文字多种输入格式

使用前需配置：
1. 数据库连接 (db_config) - 必填
"""

import os
import re
from typing import Dict, Any, Optional, Union, List, Tuple
from dataclasses import dataclass, field
from collections import defaultdict


class OrderToHuadingTemplate:
    """订单转华鼎出库单模板Skill"""
    
    VERSION = "2.0"
    
    # 字段名映射表：所有可能的字段名变体 → 标准字段名
    FIELD_ALIAS_MAPPING = {
        # 门店相关
        "store_name": ["门店", "店名", "店铺名称", "店铺", "shop", "shop_name", "name", "收货方", "客户名称", "公司名称", "客户"],
        
        # 联系方式
        "phone": ["电话", "手机", "联系电话", "tel", "mobile", "号码", "联系手机", "手机号"],
        "contact_person": ["联系人", "收货人", "contact", "收货人姓名", "收件人"],
        "address": ["地址", "收货地址", "配送地址", "addr", "详细地址"],
        
        # 商品
        "product_name": ["商品", "货品", "产品名", "product", "goods", "品名", "商品名称", "名称", "货品名称"],
        "quantity": ["数量", "件数", "箱数", "qty", "num", "订单数量", "订货数量"],
        "unit": ["单位", "件", "箱", "unit", "包装", "包装单位"],
        "spec": ["规格", "spec", "规格型号", "型号"],
        "product_code": ["编码", "商品编码", "SKU", "货号", "编号"],
        "remark": ["备注", "note", "备注信息"],
        
        # 订单
        "items": ["items", "商品明细", "商品列表", "订货明细", "明细"],
        
        # 订单
        "order_no": ["订单号", "单号", "订单编号", "order", "order_no", "送货单号", "单据编号"],
    }
    
    # 标准字段名列表（用于校验）
    STANDARD_FIELDS = list(FIELD_ALIAS_MAPPING.keys())
    
    # 华鼎模板31字段
    HUADING_FIELDS = [
        "序号", "门店编号", "门店三方编码", "仓库编码", "加急程度\n（0：普通，1：加急）",
        "商品SKU编号", "商品三方SPEC编号", "单位类型", "出库数量",
        "指定库存状态", "出库类型", "配送方式", "指定车型（专配）",
        "是否垫付", "付款方式", "快递公司", "单价", "总金额",
        "是否制定批次", "批次号", "生产日期", "备注", "生产厂家编号",
        "门店收货地址编码", "三方单号", "业务模式", "业务类型",
        "收货人", "联系电话", "收货地址", "C端快递公司"
    ]
    
    # 有默认值的字段
    DEFAULT_VALUES = {
        "加急程度": 0,
        "指定库存状态": "正常",
        "出库类型": 201,
        "配送方式": "共配",
        "是否垫付": "否",
        "是否制定批次": "否"
    }
    
    # ========== 字段名标准化相关方法 ==========
    
    @staticmethod
    def find_standard_field(input_name: str) -> Optional[str]:
        """
        将任意字段名映射到标准字段名
        
        Args:
            input_name: 原始字段名
            
        Returns:
            标准字段名，或None如果未找到
        """
        if not input_name:
            return None
        
        input_name = str(input_name).strip().lower()
        
        for std_field, aliases in OrderToHuadingTemplate.FIELD_ALIAS_MAPPING.items():
            # 完全匹配（不区分大小写）
            if input_name == std_field.lower():
                return std_field
            
            # 别名匹配
            for alias in aliases:
                if input_name == alias.lower():
                    return std_field
        
        return None
    
    def normalize_ai_result(self, raw_data: Dict) -> Tuple[Dict, List[str]]:
        """
        AI返回的字段名标准化
        
        Args:
            raw_data: AI返回的原始数据 dict
            
        Returns:
            (标准化后的数据, 警告信息列表)
        """
        normalized = {}
        warnings = []
        unknown_fields = []
        
        for ai_field_name, value in raw_data.items():
            # 跳过特殊字段
            if ai_field_name.startswith("__") or ai_field_name in ["raw_text"]:
                continue
            
            std_field = self.find_standard_field(ai_field_name)
            
            if std_field:
                normalized[std_field] = value
                if std_field != ai_field_name and ai_field_name.lower() != std_field.lower():
                    warnings.append(f"字段名'{ai_field_name}'→'{std_field}'")
            else:
                unknown_fields.append(ai_field_name)
        
        if unknown_fields:
            warnings.append(f"未知字段名{len(unknown_fields)}个，已忽略: {', '.join(unknown_fields[:5])}")
        
        return normalized, warnings
    
    def normalize_excel_headers(self, df_raw, header_row_hints: List[int] = None) -> Tuple[Dict[int, str], List[str]]:
        """
        Excel按表头名读取，找到表头行并建立列索引→标准字段名的映射
        
        Args:
            df_raw: pandas DataFrame
            header_row_hints: 可能的表头行索引列表，默认[3, 4, 5, 6]
            
        Returns:
            (列索引→标准字段名映射, 警告信息列表)
        """
        warnings = []
        header_row_hints = header_row_hints or [3, 4, 5, 6]
        
        col_mapping = {}  # {列索引: 标准字段名}
        
        for row_idx in header_row_hints:
            if row_idx >= len(df_raw):
                continue
            
            row = df_raw.iloc[row_idx]
            temp_mapping = {}
            matched_count = 0
            
            for col_idx, cell in enumerate(row):
                if cell is None or (hasattr(cell, '__iter__') and not isinstance(cell, str)):
                    if pd.isna(cell) if hasattr(pd, 'isna') else not cell:
                        continue
                
                cell_str = str(cell).strip() if cell else ""
                if not cell_str:
                    continue
                
                std_field = self.find_standard_field(cell_str)
                if std_field:
                    temp_mapping[col_idx] = std_field
                    matched_count += 1
            
            # 选择匹配数最多的行作为表头行
            if matched_count > len(col_mapping):
                col_mapping = temp_mapping
                if matched_count > 0:
                    warnings.append(f"表头行推断: 第{row_idx + 1}行，匹配{len(col_mapping)}个字段")
            
            # 如果已经匹配到关键字段（store_name和items相关），可以提前结束
            if all(v in col_mapping.values() for v in ["store_name", "product_name", "quantity"]):
                break
        
        if not col_mapping:
            warnings.append("未能识别Excel表头行，将使用默认列映射")
        
        return col_mapping, warnings
    
    def normalize_text_fields(self, text: str) -> Tuple[Dict, List[str]]:
        """
        文本按字段名映射表匹配
        支持多种格式：字段名：值、字段名 值
        
        Args:
            text: 原始文本
            
        Returns:
            (字段→值映射, 警告信息列表)
        """
        normalized = {}
        warnings = []
        
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        
        for line in lines:
            # 跳过空行
            if not line:
                continue
            
            matched = False
            
            for std_field, aliases in self.FIELD_ALIAS_MAPPING.items():
                # 跳过items的处理
                if std_field == "items":
                    continue
                
                for alias in aliases:
                    # 检查字段名是否在行首
                    if line.startswith(alias):
                        # 截取字段名之后的内容
                        value = line[len(alias):]
                        # 去除开头的标点符号和空格（：:,，、等）
                        value = re.sub(r'^[\s：:,，、]+', '', value)
                        
                        if std_field not in normalized and value:
                            normalized[std_field] = value
                            warnings.append(f"文本字段'{alias}'→'{std_field}'")
                            matched = True
                        
                        if matched:
                            break
                if matched:
                    break
        
        return normalized, warnings
    
    def validate_normalized_data(self, data: Dict) -> Dict[str, Any]:
        """
        校验标准化后的数据
        
        Args:
            data: 标准化后的数据 dict
            
        Returns:
            {
                "is_valid": bool,
                "warnings": List[str],
                "errors": List[str],
                "suggestions": List[str]
            }
        """
        result = {
            "is_valid": True,
            "warnings": [],
            "errors": [],
            "suggestions": []
        }
        
        # 1. 检查必要字段
        if not data.get("store_name"):
            result["errors"].append("缺少门店名字段")
            result["is_valid"] = False
        
        if not data.get("items") or len(data.get("items", [])) == 0:
            # 检查items是否是dict列表或需要进一步处理
            items = data.get("items", [])
            if isinstance(items, list) and len(items) == 0:
                result["warnings"].append("商品明细为空")
            elif not isinstance(items, list):
                result["errors"].append("商品明细格式错误")
                result["is_valid"] = False
        
        # 2. 检查值格式
        phone = data.get("phone", "")
        if phone:
            # 简单手机号格式校验（中国大陆）
            phone_clean = re.sub(r'\\D', '', phone)
            if len(phone_clean) > 0 and (len(phone_clean) < 7 or len(phone_clean) > 13):
                result["warnings"].append(f"电话号码格式可疑: {phone}")
        
        # 3. 检查数量
        items = data.get("items", [])
        if isinstance(items, list):
            for idx, item in enumerate(items):
                if isinstance(item, dict):
                    qty = item.get("quantity", 0)
                    if qty and (not isinstance(qty, (int, float)) or qty < 0):
                        result["errors"].append(f"商品{idx + 1}数量异常: {qty}")
                        result["is_valid"] = False
        
        return result
    
    def __init__(
        self,
        db_config: Optional[Dict] = None,
        output_dir: Optional[str] = None
    ):
        """
        初始化Skill配置
        
        Args:
            db_config: 数据库连接配置（必填）
            output_dir: 输出目录（可选，默认./output）
        """
        if not db_config:
            raise ValueError("db_config 是必填参数")
        
        self.shipper_id = None  # 不再直接传入，通过门店匹配获取
        self.db_config = db_config
        self.output_dir = output_dir or "./output"
        
        # 从数据库加载仓库编码映射
        self.warehouse_code_map = self._load_warehouse_mapping()
        
        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _load_warehouse_mapping(self) -> Dict[str, str]:
        """从数据库加载仓库编码映射"""
        try:
            import psycopg2
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            cur.execute("SELECT warehouse_name, warehouse_code FROM warehouse_code_mapping")
            rows = cur.fetchall()
            
            conn.close()
            
            mapping = {}
            for name, code in rows:
                mapping[name] = code
            
            return mapping
            
        except Exception as e:
            print(f"加载仓库编码映射失败: {e}")
            return {}
    
    def _get_warehouse_code(self, warehouse_name: str) -> str:
        """获取仓库编码，找不到则报错"""
        if not warehouse_name:
            raise ValueError(f"门店的仓库名称为空，请检查门店数据")
        
        warehouse_code = self.warehouse_code_map.get(warehouse_name)
        if not warehouse_code:
            available = ", ".join(self.warehouse_code_map.keys())
            raise ValueError(
                f"仓库'{warehouse_name}'在仓库编码映射表中未找到！\n"
                f"请检查仓库编码表，确保该仓库已配置。\n"
                f"当前可用的仓库：{available}"
            )
        
        return warehouse_code
    
    def _detect_input_type(self, order_input: str) -> str:
        """
        自动检测输入类型
        
        Returns:
            "excel" / "image" / "pdf" / "text"
        """
        if not order_input:
            return "text"
        
        # 1. 检查是否为文件路径
        if os.path.exists(order_input):
            ext = os.path.splitext(order_input)[1].lower()
            if ext in [".xlsx", ".xls"]:
                return "excel"
            elif ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif"]:
                return "image"
            elif ext == ".pdf":
                return "pdf"
            else:
                # 未知扩展名，尝试作为图片处理
                return "image"
        
        # 2. 检查扩展名字符串
        input_str = str(order_input).strip().lower()
        if input_str.endswith(".xlsx") or input_str.endswith(".xls"):
            return "excel"
        elif input_str.endswith(".jpg") or input_str.endswith(".png") or input_str.endswith(".jpeg"):
            return "image"
        elif input_str.endswith(".pdf"):
            return "pdf"
        
        # 3. 纯文本判断（包含换行或结构化标记）
        if "\n" in order_input or "订单号" in order_input or "门店" in order_input:
            return "text"
        
        # 4. 默认当作图片路径处理（网络URL或相对路径）
        return "image"
    
    def _parse_image(self, image_path: str) -> Optional[Dict]:
        """
        图片解析 - 返回需要外部OCR的标记
        
        由于图像识别需要AI能力，此方法返回特殊状态码 "NEED_OCR"
        由调用方使用image工具进行识别后，将结果传入_handle_ocr_result()
        
        Args:
            image_path: 图片文件路径
            
        Returns:
            {"__need_ocr__: True, "image_path": image_path} 或 None
        """
        if not image_path or not os.path.exists(image_path):
            return None
        
        # 返回需要OCR的标记
        return {"__need_ocr__": True, "image_path": image_path}
    
    def _handle_ocr_result(self, ocr_result: Dict) -> Optional[Dict]:
        """
        处理OCR识别结果，将结构化数据规范化
        
        Args:
            ocr_result: OCR工具返回的结构化数据
            
        Returns:
            规范化后的订单数据字典
        """
        if not ocr_result or not isinstance(ocr_result, dict):
            return None
        
        # 如果OCR结果已经是结构化数据，直接规范化
        return self._normalize_extracted_data(ocr_result)
    
    def _parse_pdf(self, pdf_path: str) -> Optional[Dict]:
        """
        PDF解析 - 返回需要外部OCR的标记
        
        由于PDF识别需要AI能力，此方法返回特殊状态码 "__need_pdf_ocr__"
        由调用方使用pdf工具进行识别后，将结果传入_handle_pdf_ocr_result()
        
        Args:
            pdf_path: PDF文件路径
            
        Returns:
            {"__need_pdf_ocr__": True, "pdf_path": pdf_path} 或 None
        """
        if not pdf_path or not os.path.exists(pdf_path):
            return None
        
        # 返回需要PDF OCR的标记
        return {"__need_pdf_ocr__": True, "pdf_path": pdf_path}
    
    def _handle_pdf_ocr_result(self, pdf_ocr_result: Dict) -> Optional[Dict]:
        """
        处理PDF OCR识别结果，将结构化数据规范化
        
        Args:
            pdf_ocr_result: PDF工具返回的结构化数据
            
        Returns:
            规范化后的订单数据字典
        """
        if not pdf_ocr_result or not isinstance(pdf_ocr_result, dict):
            return None
        
        # 如果OCR结果已经是结构化数据，直接规范化
        return self._normalize_extracted_data(pdf_ocr_result)
    
    def _parse_text(self, text: str) -> Optional[Dict]:
        """
        从纯文本/粘贴内容中解析订单数据
        支持多种常见格式
        """
        return self._normalize_extracted_data({"raw_text": text})
    
    def _normalize_extracted_data(self, extracted: Dict) -> Optional[Dict]:
        """
        规范化从OCR/文本提取的数据，适配标准订单格式
        先进行字段名标准化，再处理数据
        """
        raw_text = extracted.get("raw_text", "")
        
        # 如果是纯文本，先解析
        if raw_text:
            return self._parse_raw_text(raw_text)
        
        # ========== Step 1: 字段名标准化 ==========
        normalized_fields, norm_warnings = self.normalize_ai_result(extracted)
        
        # 如果标准化后没有store_name，可能是识别问题
        if not normalized_fields.get("store_name") and not extracted.get("items"):
            # 尝试原始数据的keys
            if not normalized_fields:
                normalized_fields = {k: v for k, v in extracted.items() 
                                   if not k.startswith("__") and k != "raw_text"}
        
        # ========== Step 2: 提取各字段 ==========
        order_no = str(normalized_fields.get("order_no", extracted.get("order_no", ""))).strip()
        store_name = str(normalized_fields.get("store_name", extracted.get("store_name", ""))).strip()
        
        # 清理门店名称
        for prefix in ['河北-', '天津-', '沧州-']:
            if store_name.startswith(prefix):
                store_name = store_name[len(prefix):]
                break
        
        # 提取商品明细
        items_raw = normalized_fields.get("items", extracted.get("items", []))
        items = []
        for idx, item in enumerate(items_raw if isinstance(items_raw, list) else [], start=1):
            if not isinstance(item, dict):
                continue
            
            # 支持商品名字段名变体
            product_name = None
            for pn_key in ["product_name", "product", "goods", "name", "品名"]:
                if pn_key in item:
                    product_name = str(item.get(pn_key, "")).strip()
                    break
            
            if not product_name or product_name == "nan":
                continue
            
            # 数量处理
            qty = None
            for qty_key in ["quantity", "qty", "num", "数量"]:
                if qty_key in item:
                    qty = item.get(qty_key, 0)
                    break
            if qty is None:
                qty = 0
            if isinstance(qty, str):
                qty = int(float(re.sub(r'[^0-9.]', '', qty)))
            qty = int(qty) if qty else 0
            
            items.append({
                "seq": item.get("seq", idx),
                "product_code": str(item.get("product_code", item.get("code", ""))).strip(),
                "product_name": product_name,
                "spec": str(item.get("spec", item.get("规格", ""))).strip(),
                "quantity": qty,
                "unit": str(item.get("unit", item.get("单位", "件"))).strip(),
                "remark": str(item.get("remark", item.get("note", ""))).strip()
            })
        
        if not store_name and not items:
            return None
        
        return {
            "order_no": order_no,
            "store_name": store_name,
            "contact_person": str(normalized_fields.get("contact_person", extracted.get("contact_person", ""))).strip(),
            "phone": str(normalized_fields.get("phone", extracted.get("phone", ""))).strip(),
            "address": str(normalized_fields.get("address", extracted.get("address", ""))).strip(),
            "items": items,
            "_norm_warnings": norm_warnings  # 保留警告信息供后续使用
        }
    
    def _parse_raw_text(self, text: str) -> Optional[Dict]:
        """
        解析原始粘贴文本，使用字段名映射表匹配
        支持多种客户格式的字段名
        
        匹配方式：
        - 遍历每一行
        - 用字段名映射表查找这一行是否包含某个字段名
        - 如果包含，截取字段名后面的内容作为值（自动去除标点符号和空格）
        """
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        warnings = []
        
        order_no = ""
        store_name = ""
        contact_person = ""
        phone = ""
        address = ""
        items = []
        
        # 用于记录哪些字段已匹配
        matched_fields = set()
        
        for line in lines:
            # 跳过空行和注释
            if not line or line.startswith("#"):
                continue
            
            # 检查这一行是否是字段行
            line_matched = False
            
            for std_field, aliases in self.FIELD_ALIAS_MAPPING.items():
                # 跳过items的处理（items是列表，需要特殊处理）
                if std_field == "items":
                    continue
                    
                for alias in aliases:
                    # 查找字段名在行中的位置（行首）
                    if line.startswith(alias):
                        # 截取字段名之后的内容
                        value = line[len(alias):]
                        # 去除开头的标点符号和空格（：:,，、等）
                        value = re.sub(r'^[\s：:,，、]+', '', value)
                        
                        if std_field == "store_name" and not store_name:
                            store_name = value
                            matched_fields.add("store_name")
                            warnings.append(f"文本字段'{alias}'→'store_name'")
                            line_matched = True
                        elif std_field == "order_no" and not order_no:
                            order_no = value
                            matched_fields.add("order_no")
                            line_matched = True
                        elif std_field == "contact_person" and not contact_person:
                            contact_person = value
                            matched_fields.add("contact_person")
                            line_matched = True
                        elif std_field == "phone" and not phone:
                            phone = value
                            matched_fields.add("phone")
                            line_matched = True
                        elif std_field == "address" and not address:
                            address = value
                            matched_fields.add("address")
                            line_matched = True
                        
                        if line_matched:
                            break
                if line_matched:
                    break
            
            if line_matched:
                continue
            
            # ========== 商品明细识别 ==========
            # 商品明细识别（以 -、*、序号 开头）
            item_match = re.match(r'^[\-\*\d]+[\.、\s]+(.+?)(?:\s+(\d+)[\u4e00-\u9fa5件个箱袋台]?)?$', line)
            if item_match:
                product_name = item_match.group(1).strip()
                quantity = int(item_match.group(2)) if item_match.group(2) else 1
                items.append({
                    "seq": len(items) + 1,
                    "product_code": "",
                    "product_name": product_name,
                    "spec": "",
                    "quantity": quantity,
                    "unit": "件",
                    "remark": ""
                })
                continue
            
            # 如果这行只有中文/英文/数字，且较长，可能是门店名
            if re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9]{4,}$', line):
                if not store_name and "store_name" not in matched_fields:
                    store_name = line
                    warnings.append(f"推断门店名: '{line}'")
        
        # 清理门店名
        for prefix in ['河北-', '天津-', '沧州-']:
            if store_name.startswith(prefix):
                store_name = store_name[len(prefix):]
                break
        
        if not store_name and not items:
            warnings.append("未能从文本中提取到门店名和商品明细")
        
        return {
            "order_no": order_no,
            "store_name": store_name,
            "contact_person": contact_person,
            "phone": phone,
            "address": address,
            "items": items,
            "_norm_warnings": warnings
        }
        
        # 如果这行只有中文/英文/数字，且较长，可能是门店名
        if re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9]{4,}$', line):
            if not store_name and "store_name" not in matched_fields:
                store_name = line
                warnings.append(f"推断门店名: '{line}'")
        
        # 清理门店名
        for prefix in ['河北-', '天津-', '沧州-']:
            if store_name.startswith(prefix):
                store_name = store_name[len(prefix):]
                break
        
        if not store_name and not items:
            warnings.append("未能从文本中提取到门店名和商品明细")
        
        return {
            "order_no": order_no,
            "store_name": store_name,
            "contact_person": contact_person,
            "phone": phone,
            "address": address,
            "items": items,
            "_norm_warnings": warnings
        }
    
    def execute(self, order_input: str = None, output_file: str = None, order_type: str = "auto", 
                ocr_result: Dict = None) -> Dict[str, Any]:
        """
        执行订单转华鼎模板（支持多格式输入）
        
        Args:
            order_input: 订单输入（文件路径/图片路径/文字内容）
            output_file: 输出文件路径
            order_type: 输入类型（"auto"/"excel"/"image"/"pdf"/"text"）
            ocr_result: 图片/PDF OCR识别结果（当图片/PDF解析返回NEED_OCR时，由调用方识别后传入）
        
        Returns:
            Dict包含:
            - success: 是否成功
            - need_ocr: 是否需要OCR识别（当图片无法直接处理时返回True）
            - output_file: 输出文件路径
            - order_no: 订单号
            - store_code: 门店编号
            - warehouse_code: 仓库编码
            - item_count: 商品数量
            - unmatched_items: 未匹配SKU的商品列表
            - extracted_from: 输入来源（excel/image/pdf/text）
            - message: 消息
        """
        try:
            # 自动检测类型
            if order_type == "auto":
                order_type = self._detect_input_type(order_input)
            
            # 解析订单数据
            if order_type == "image":
                if ocr_result:
                    # 有OCR结果，直接处理
                    order_data = self._handle_ocr_result(ocr_result)
                    extracted_from = "image"
                elif order_input:
                    # 没有OCR结果，检查图片文件
                    parse_result = self._parse_image(order_input)
                    if parse_result and parse_result.get("__need_ocr__"):
                        # 返回需要OCR的标记
                        return {
                            "success": False,
                            "need_ocr": True,
                            "image_path": order_input,
                            "message": f"图片需要OCR识别，请稍候"
                        }
                    else:
                        order_data = parse_result
                        extracted_from = "image"
                else:
                    order_data = None
                    extracted_from = "image"
            elif order_type == "pdf":
                if ocr_result:
                    # 有OCR结果，直接处理
                    order_data = self._handle_pdf_ocr_result(ocr_result)
                    extracted_from = "pdf"
                elif order_input:
                    # 没有OCR结果，检查PDF文件
                    parse_result = self._parse_pdf(order_input)
                    if parse_result and parse_result.get("__need_pdf_ocr__"):
                        # 返回需要PDF OCR的标记
                        return {
                            "success": False,
                            "need_pdf_ocr": True,
                            "pdf_path": order_input,
                            "message": f"PDF需要OCR识别，请稍候"
                        }
                    else:
                        order_data = parse_result
                        extracted_from = "pdf"
                else:
                    order_data = None
                    extracted_from = "pdf"
            elif order_type == "text":
                order_data = self._parse_text(order_input)
                extracted_from = "text"
            else:
                # excel
                if not os.path.exists(order_input):
                    return {"success": False, "message": f"订单文件不存在: {order_input}"}
                order_data = self._parse_order_excel(order_input)
                extracted_from = "excel"
            
            if not order_data:
                return {"success": False, "message": f"订单解析失败（来源：{order_type}），请检查格式或上传Excel文件"}
            
            store_info = self._match_store(order_data["store_name"])
            
            # 门店未匹配到，返回错误
            if not store_info:
                return {
                    "success": False,
                    "need_store_match": True,
                    "store_name_submitted": order_data.get("store_name", ""),
                    "message": f"门店「{order_data.get('store_name', '未知')}」未找到匹配门店，请确认门店是否在数据库中"
                }
            
            # 模糊匹配返回候选列表，需要用户确认
            if store_info.get("need_confirm"):
                return {
                    "success": False,
                    "need_store_confirm": True,
                    "store_name_submitted": store_info.get("store_name_submitted", ""),
                    "candidates": store_info.get("candidates", []),
                    "message": f"门店「{store_info.get('store_name_submitted', '未知')}」找到 {len(store_info.get('candidates', []))} 个候选门店，请选择或确认"
                }
            
            # 门店匹配失败但有相似货主提示
            if store_info.get("need_customer_hint"):
                return {
                    "success": False,
                    "need_customer_hint": True,
                    "store_name_submitted": store_info.get("store_name_submitted", ""),
                    "possible_customers": store_info.get("possible_customers", []),
                    "message": f"门店「{store_info.get('store_name_submitted', '未知')}」未找到匹配门店，但可能属于以下货主："
                }
            
            owner_code = store_info.get("owner_code", self.shipper_id) if store_info else self.shipper_id
            sku_results, unmatched_items = self._match_sku(order_data["items"], owner_code)
            
            if not output_file:
                order_no_safe = order_data["order_no"].replace("/", "-").replace("\\", "-") if order_data.get("order_no") else "unknown"
                output_file = os.path.join(self.output_dir, f"华鼎出库单_{order_no_safe}.xlsx")
            
            self._generate_template(order_data, store_info, sku_results, output_file)
            
            # 构建返回消息
            message = "模板生成成功"
            if unmatched_items:
                unmatched_details = []
                for item in unmatched_items:
                    details = f"序号{item['seq']}: {item['product_name']} | 编码:{item['product_code']} | 规格:{item['spec']} | 数量:{item['quantity']}件"
                    unmatched_details.append(details)
                message = f"模板生成成功，但有{len(unmatched_items)}条商品未匹配SKU：" + "; ".join(unmatched_details)
            
            return {
                "success": True,
                "output_file": output_file,
                "order_no": order_data.get("order_no", ""),
                "store_code": store_info.get("store_code", "") if store_info else "",
                "owner_code": owner_code,
                "warehouse_code": self._get_warehouse_code(store_info.get("warehouse_name", "")) if store_info else "",
                "item_count": len(sku_results),
                "unmatched_count": len(unmatched_items),
                "unmatched_items": unmatched_items,
                "extracted_from": extracted_from,
                "message": message
            }
            
        except Exception as e:
            return {"success": False, "message": f"错误: {str(e)}"}
    
    def _parse_order_excel(self, order_file: str) -> Optional[Dict]:
        """
        解析客户订单Excel - 按表头名读取，而非固定行列
        
        支持不同客户的不同Excel格式
        """
        try:
            import pandas as pd
            
            df_raw = pd.read_excel(order_file, header=None)
            warnings = []
            
            # ========== Step 1: 找到表头行，建立列映射 ==========
            col_mapping, header_warnings = self.normalize_excel_headers(df_raw)
            warnings.extend(header_warnings)
            
            # ========== Step 2: 提取元数据（门店、订单号等）============
            # 尝试从常见位置读取元数据（兼容旧格式）
            order_no = ""
            store_name = ""
            contact_person = ""
            phone = ""
            address = ""
            
            # 尝试从col_mapping读取
            if "order_no" in col_mapping:
                col_idx = col_mapping["order_no"]
                if len(df_raw) > 1:
                    order_no = str(df_raw.iloc[1, col_idx]) if pd.notna(df_raw.iloc[1, col_idx]) else ""
            
            if "store_name" in col_mapping:
                col_idx = col_mapping["store_name"]
                if len(df_raw) > 2:
                    store_name = str(df_raw.iloc[2, col_idx]) if pd.notna(df_raw.iloc[2, col_idx]) else ""
            
            if "contact_person" in col_mapping:
                col_idx = col_mapping["contact_person"]
                if len(df_raw) > 2:
                    contact_person = str(df_raw.iloc[2, col_idx]) if pd.notna(df_raw.iloc[2, col_idx]) else ""
            
            if "phone" in col_mapping:
                col_idx = col_mapping["phone"]
                if len(df_raw) > 3:
                    phone = str(df_raw.iloc[3, col_idx]) if pd.notna(df_raw.iloc[3, col_idx]) else ""
            
            # Fallback: 如果col_mapping中没有store_name等字段，直接从固定位置读取
            if not store_name and len(df_raw) > 2 and len(df_raw.columns) > 1:
                # 门店名通常在 row=2, col=1
                val = df_raw.iloc[2, 1]
                if pd.notna(val):
                    store_name = str(val).strip()
            
            if not order_no and len(df_raw) > 1 and len(df_raw.columns) > 1:
                # 订单号通常在 row=1, col=1
                val = df_raw.iloc[1, 1]
                if pd.notna(val):
                    order_no = str(val).strip()
            
            if not contact_person and len(df_raw) > 2:
                # 联系人通常在 row=2, col=6
                val = df_raw.iloc[2, 6]
                if pd.notna(val):
                    contact_person = str(val).strip()
            
            if not phone and len(df_raw) > 3:
                # 电话通常在 row=3, col=6
                val = df_raw.iloc[3, 6]
                if pd.notna(val):
                    phone = str(val).strip()
            
            # 清理门店名称（去掉前缀）
            for prefix in ['河北-', '天津-', '沧州-']:
                if store_name.startswith(prefix):
                    store_name = store_name.replace(prefix, "")
                    break
            
            # ========== Step 3: 提取商品明细 ==========
            items = []
            data_start_row = 6  # 默认数据开始行
            
            # 找到表头行后，确定数据开始行
            if col_mapping:
                # 找到表头行的位置
                for row_idx in [3, 4, 5, 6]:
                    if row_idx >= len(df_raw):
                        continue
                    row = df_raw.iloc[row_idx]
                    has_header = any(
                        pd.notna(cell) and self.find_standard_field(str(cell))
                        for cell in row
                    )
                    if has_header:
                        data_start_row = row_idx + 1
                        break
            
            for i in range(data_start_row, len(df_raw)):
                seq = df_raw.iloc[i, 0] if pd.notna(df_raw.iloc[i, 0]) else None
                if seq is None or str(seq) == "合计：" or str(seq).startswith("合计"):
                    continue
                
                # 尝试多种列位置读取商品信息
                product_code = ""
                product_name = ""
                spec = ""
                quantity = 0
                unit = "件"
                remark = ""
                
                # 商品编码：col 1
                if pd.notna(df_raw.iloc[i, 1]):
                    product_code = str(df_raw.iloc[i, 1]).strip()
                
                # 商品名称：col 2 或 col 1
                if pd.notna(df_raw.iloc[i, 2]):
                    product_name = str(df_raw.iloc[i, 2]).strip()
                elif pd.notna(df_raw.iloc[i, 1]) and not product_code:
                    # 如果col1没有编码，可能是商品名称
                    potential_name = str(df_raw.iloc[i, 1]).strip()
                    if potential_name and potential_name != "nan" and not potential_name.isdigit():
                        product_name = potential_name
                        product_code = ""
                
                if not product_name or product_name == "nan":
                    continue
                
                # 规格：col 3
                if pd.notna(df_raw.iloc[i, 3]):
                    spec = str(df_raw.iloc[i, 3]).replace("规格：", "").replace("件：", "").strip()
                
                # 数量：col 5
                if pd.notna(df_raw.iloc[i, 5]):
                    qty_val = df_raw.iloc[i, 5]
                    quantity = int(qty_val) if isinstance(qty_val, (int, float)) else 0
                
                # 单位：col 6
                if pd.notna(df_raw.iloc[i, 6]):
                    unit = str(df_raw.iloc[i, 6]).strip()
                
                # 备注：col 8
                if pd.notna(df_raw.iloc[i, 8]):
                    remark = str(df_raw.iloc[i, 8]).strip()
                
                items.append({
                    "seq": int(seq) if isinstance(seq, (int, float)) else len(items) + 1,
                    "product_code": product_code,
                    "product_name": product_name,
                    "spec": spec,
                    "quantity": quantity,
                    "unit": unit,
                    "remark": remark
                })
            
            if not store_name and not items:
                warnings.append("未能从Excel中提取到门店名和商品明细")
            
            return {
                "order_no": order_no,
                "store_name": store_name,
                "contact_person": contact_person,
                "phone": phone,
                "address": address,
                "items": items,
                "_norm_warnings": warnings
            }
            
        except Exception as e:
            print(f"订单解析错误: {e}")
            return None
    
    def _match_store(self, store_name: str) -> Optional[Dict]:
        """
        门店匹配 - 两层匹配 + 货主提示策略
        
        第1层：精确匹配（store_name = 输入名称）
        第2层：模糊匹配（相似度计算），相似度<75%视为匹配不到
        第3层：门店匹配失败时，搜索customer表查找相似货主名
        
        Returns:
            精确匹配时：返回门店信息 dict
            模糊匹配时：返回 {"need_confirm": True, "candidates": [...], "store_name_submitted": ...}
            匹配不到但有相似货主：返回 {"need_customer_hint": True, "possible_customers": [...]}
            匹配不到时：返回 None
        """
        try:
            import psycopg2
            from difflib import SequenceMatcher
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            def calc_similarity(s1, s2):
                """计算两个字符串的相似度 (0-1)
                
                策略：
                1. 先做整体相似度（SequenceMatcher）
                2. 如果短串是长串的前缀/包含，额外加分
                3. 清理品牌前缀后再比较
                4. 核心关键词匹配（地名/店名等）加分
                """
                if not s1 or not s2:
                    return 0.0
                
                # 品牌前缀列表（按长度降序排列，避免误匹配）
                prefixes = ['制茶青年-', '制茶青年', '天津仓-', '天津仓']
                
                # 清理品牌前缀
                s1_cleaned = s1
                s2_cleaned = s2
                for prefix in prefixes:
                    if s1_cleaned.startswith(prefix):
                        s1_cleaned = s1_cleaned[len(prefix):]
                    if s2_cleaned.startswith(prefix):
                        s2_cleaned = s2_cleaned[len(prefix):]
                
                # 1. 基础相似度
                base_sim = SequenceMatcher(None, s1_cleaned, s2_cleaned).ratio()
                
                # 2. 包含匹配加分（短串是长串的前缀/包含）
                # 确保s1是较短的串
                shorter = s1_cleaned if len(s1_cleaned) <= len(s2_cleaned) else s2_cleaned
                longer = s2_cleaned if len(s1_cleaned) <= len(s2_cleaned) else s1_cleaned
                
                contain_bonus = 0.0
                if shorter and shorter in longer:
                    # 包含匹配相似度 = 短串长度/长串长度 + 0.3
                    contain_bonus = len(shorter) / len(longer) + 0.3
                
                # 3. 核心关键词匹配（地名关键词）
                # 提取关键位置词：省/市/区/县/镇/乡/村/店等
                location_pattern = r'[\u4e00-\u9fa5]{2,6}(?:省|市|区|县|镇|乡|村|街|路|店|仓|口)'
                import re
                loc1 = set(re.findall(location_pattern, s1_cleaned))
                loc2 = set(re.findall(location_pattern, s2_cleaned))
                
                keyword_bonus = 0.0
                if loc1 and loc2:
                    # 有共同关键词则加分
                    common = loc1 & loc2
                    if common:
                        # 每个共同关键词加0.2，最高0.6
                        keyword_bonus = min(len(common) * 0.2, 0.6)
                
                return max(base_sim, contain_bonus, base_sim + keyword_bonus)
            
            def try_exact_match(name):
                """精确匹配：store_name = 输入名称"""
                cur.execute("""
                    SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                    FROM store_list
                    WHERE store_name = %s
                    LIMIT 1
                """, (name,))
                return cur.fetchone()
            
            def try_fuzzy_match(name, top_n=5):
                """模糊匹配：返回多个候选，按相似度排序
                
                策略：
                1. 先用完整名称LIKE查询
                2. 如果没有结果，拆分为关键词逐个查询后合并去重
                """
                all_rows = []
                
                # 第1步：完整名称LIKE
                cur.execute("""
                    SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                    FROM store_list
                    WHERE store_name LIKE %s
                    LIMIT 50
                """, (f"%{name}%",))
                all_rows = list(cur.fetchall())
                
                # 第2步：如果没有结果，提取关键词分别查询
                if not all_rows:
                    # 提取2-4个字的中文关键词
                    import re
                    keywords = re.findall(r'[\u4e00-\u9fa5]{2,4}', name)
                    seen_codes = set()
                    for kw in keywords:
                        if len(kw) >= 2:
                            cur.execute("""
                                SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                                FROM store_list
                                WHERE store_name LIKE %s
                                LIMIT 20
                            """, (f"%{kw}%",))
                            for row in cur.fetchall():
                                if row[0] not in seen_codes:
                                    all_rows.append(row)
                                    seen_codes.add(row[0])
                
                # 计算相似度并排序
                scored = []
                for row in all_rows:
                    db_name = row[1] or ""
                    sim = calc_similarity(name, db_name)
                    scored.append((sim, row))
                
                scored.sort(key=lambda x: x[0], reverse=True)
                return scored[:top_n]
            
            def search_similar_customers(name):
                """搜索customer表查找相似货主名"""
                cur.execute("""
                    SELECT customer_id, customer_name, warehouse_name, status
                    FROM customer
                    WHERE status = 'ACTIVE'
                    LIMIT 20
                """)
                rows = cur.fetchall()
                
                # 计算相似度
                scored = []
                for row in rows:
                    cust_name = row[1] or ""
                    # 相似度：门店名 vs 货主名
                    sim = calc_similarity(name, cust_name)
                    scored.append((sim, row))
                
                scored.sort(key=lambda x: x[0], reverse=True)
                # 只返回相似度>=75%的
                return [(sim, row) for sim, row in scored if sim >= 0.70]
            
            # ========== 第1层：精确匹配 ==========
            row = try_exact_match(store_name)
            
            if row:
                # 精确匹配成功，直接返回
                warehouse_name = row[2] or ""
                conn.close()
                return {
                    "store_code": row[0] or "",
                    "store_name": row[1] or "",
                    "warehouse_name": warehouse_name,
                    "warehouse_code": self._get_warehouse_code(warehouse_name),
                    "address": row[3] or "",
                    "contact_person": row[4] or "",
                    "phone": row[5] or "",
                    "owner_code": row[6] or "",
                    "match_type": "exact",
                    "match_method": "精确匹配"
                }
            
            # ========== 第2层：模糊匹配（仅在精确匹配失败时）==========
            candidates = try_fuzzy_match(store_name)
            
            # 过滤：相似度 < 75% 的结果视为匹配不到
            high_confidence = [(sim, row) for sim, row in candidates if sim >= 0.70]
            
            if high_confidence:
                # 返回相似度 >= 75% 的候选列表（需要用户确认）
                conn.close()
                
                candidate_list = []
                for sim, row in high_confidence:
                    warehouse_name = row[2] or ""
                    candidate_list.append({
                        "store_code": row[0] or "",
                        "store_name": row[1] or "",
                        "warehouse_name": warehouse_name,
                        "warehouse_code": self._get_warehouse_code(warehouse_name),
                        "address": row[3] or "",
                        "contact_person": row[4] or "",
                        "phone": row[5] or "",
                        "owner_code": row[6] or "",
                        "similarity": round(sim * 100, 1)
                    })
                
                return {
                    "need_confirm": True,
                    "candidates": candidate_list,
                    "store_name_submitted": store_name,
                    "match_type": "fuzzy",
                    "match_method": f"模糊匹配({len(candidate_list)}个候选)"
                }
            
            # ========== 第3层：门店匹配失败，搜索相似货主 ==========
            similar_customers = search_similar_customers(store_name)
            
            conn.close()
            
            if similar_customers:
                customer_list = []
                for sim, row in similar_customers:
                    customer_list.append({
                        "customer_id": row[0] or "",
                        "customer_name": row[1] or "",
                        "warehouse_name": row[2] or "",
                        "similarity": round(sim * 100, 1)
                    })
                
                return {
                    "need_customer_hint": True,
                    "possible_customers": customer_list,
                    "store_name_submitted": store_name,
                    "match_type": "customer_hint",
                    "match_method": f"货主提示({len(customer_list)}个候选)"
                }
            
            # 真的匹配不到
            return None
            
        except Exception as e:
            print(f"门店匹配错误: {e}")
            return None

    def _clean_product_name(self, name: str) -> str:
        """
        清理商品名称，去除特殊字符用于匹配
        
        处理规则：
        - 去除首尾特殊字符（-、/、.等）
        - 去除括号及其内容
        - 去除空格
        """
        import re
        # 去除首尾的特殊字符
        name = name.strip().rstrip('-').rstrip('/').rstrip('.')
        # 去除括号及其内容
        name = re.sub(r'[（(].*[)）]', '', name)
        # 去除空格
        name = name.strip()
        return name
    
    def _match_sku(self, items: list, owner_code: str) -> tuple:
        """
        SKU匹配 - 多层匹配策略
        
        匹配优先级：
        1. 第1层：精确匹配 - 通过 customer_sku_code（订单商品编码）在 shipper_sku_mapping 表精确匹配
        2. 第2层：名称匹配 - 用商品名称在 system_sku 表中模糊匹配
        3. 第3层：名称+规格匹配 - 在名称匹配基础上用规格进一步精确筛选
        
        Args:
            items: 订单商品列表
            owner_code: 货主ID（用于过滤门店）
            
        Returns:
            (results, unmatched_items)
        """
        try:
            import psycopg2
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            results = []
            unmatched_items = []
            
            for item in items:
                sku_code = ""
                sku_name = ""
                match_method = ""
                
                customer_sku_code = item.get("product_code", "")  # 订单中的商品编码
                product_name = item.get("product_name", "")
                spec = item.get("spec", "")
                
                # ========== 第1层：精确匹配（customer_sku_code）==========
                if customer_sku_code:
                    cur.execute("""
                        SELECT system_sku_code, system_sku_name 
                        FROM shipper_sku_mapping 
                        WHERE shipper_id = %s AND customer_sku_code = %s
                        LIMIT 1
                    """, (owner_code, customer_sku_code))
                    row = cur.fetchone()
                    if row:
                        sku_code = row[0]
                        sku_name = row[1]
                        match_method = "精确匹配"
                
                # ========== 第2层：名称匹配（SKU编码未命中时）==========
                if not sku_code and product_name:
                    # 用商品名称在 system_sku 表中模糊匹配
                    cur.execute("""
                        SELECT sku_code, sku_name, package_spec 
                        FROM system_sku 
                        WHERE sku_name LIKE %s
                    """, (f"%{product_name}%",))
                    candidates = cur.fetchall()
                    
                    if not candidates:
                        # 未匹配
                        match_method = "not_found"
                    elif len(candidates) == 1:
                        # 只有一条，直接使用
                        sku_code = candidates[0][0]
                        sku_name = candidates[0][1]
                        match_method = "名称匹配"
                    else:
                        # ========== 第3层：名称+规格匹配（多条时）==========
                        matched = None
                        for c in candidates:
                            pkg_spec = c[2] or ""
                            # 规格匹配：检查规格关键词是否匹配
                            # 例如：订单规格 "2kg*6瓶/件" 包含 "2kg"
                            spec_keywords = []
                            if spec:
                                # 提取规格中的关键单位（如 2kg、1kg、850g、50袋）
                                import re
                                spec_keywords = re.findall(r'[\d.]+(?:kg|g|袋|瓶|罐)', spec.lower())
                            
                            if spec_keywords:
                                # 检查规格关键词是否在 package_spec 中
                                if any(kw in pkg_spec.lower() for kw in spec_keywords):
                                    matched = c
                                    break
                        
                        if matched:
                            sku_code = matched[0]
                            sku_name = matched[1]
                            match_method = "名称+规格匹配"
                        else:
                            # 兜底：返回第一个
                            sku_code = candidates[0][0]
                            sku_name = candidates[0][1]
                            match_method = f"名称匹配(兜底{len(candidates)}条)"
                
                if sku_code:
                    # 查询SKU的单位类型（大/中/小单位）
                    unit_type = self._get_unit_type(sku_code, owner_code)
                    results.append({
                        "seq": item["seq"],
                        "product_name": product_name,
                        "spec": spec,
                        "quantity": item["quantity"],
                        "remark": item.get("remark", ""),
                        "sku_code": sku_code,
                        "sku_name": sku_name,
                        "unit_type": unit_type,
                        "match_method": match_method
                    })
                else:
                    unmatched_items.append({
                        "seq": item["seq"],
                        "product_code": customer_sku_code,
                        "product_name": product_name,
                        "spec": spec,
                        "quantity": item["quantity"],
                        "unit": item.get("unit", "件")
                    })
            
            conn.close()
            return results, unmatched_items
            
        except Exception as e:
            print(f"SKU匹配错误: {e}")
            return items, []
    
    def _get_unit_type(self, sku_code: str, owner_code: str) -> str:
        """
        获取SKU的单位类型（大/中/小单位）
        根据同一SKU在shipper_sku_mapping中的多个单位配置，按转换比判断
        """
        try:
            import psycopg2
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            # 查询该货主下所有SKU单位配置
            cur.execute("""
                SELECT m.unit_conversion_rule, s.unit
                FROM shipper_sku_mapping m
                JOIN system_sku s ON s.sku_code = m.system_sku_code
                WHERE m.shipper_id = %s AND m.system_sku_code = %s
            """, (owner_code, sku_code))
            rows = cur.fetchall()
            conn.close()
            
            if not rows:
                return "大单位"
            
            # 从JSONB中提取ratio
            ratios = []
            for r in rows:
                rule = r[0] if r[0] else {}
                if isinstance(rule, dict) and 'ratio' in rule:
                    ratios.append(float(rule['ratio']))
                elif isinstance(rule, str):
                    import json
                    parsed = json.loads(rule)
                    if 'ratio' in parsed:
                        ratios.append(float(parsed['ratio']))
            
            if not ratios:
                return "大单位"
            
            max_ratio = max(ratios)
            min_ratio = min(ratios)
            
            # 简单判断：如果只有一条记录，默认大单位
            if len(ratios) == 1:
                return "大单位"
            
            # 返回最大单位作为默认
            return "大单位"
                
        except Exception as e:
            print(f"获取单位类型失败: {e}")
            return "大单位"

    
    def _generate_template(self, order_data: Dict, store_info: Optional[Dict], 
                          sku_results: list, output_file: str):
        """生成华鼎模板Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "omsWare"
        
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        
        column_widths = {1: 6, 2: 16, 3: 12, 4: 8, 5: 10, 6: 18, 7: 16, 8: 10, 9: 10,
            10: 10, 11: 8, 12: 10, 13: 12, 14: 8, 15: 8, 16: 10, 17: 8, 18: 10,
            19: 10, 20: 15, 21: 12, 22: 25, 23: 12, 24: 15, 25: 18, 26: 8, 27: 8,
            28: 10, 29: 15, 30: 40, 31: 12}
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[1].height = 35
        
        cell_alignment = Alignment(horizontal="center", vertical="center")
        cell_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        warehouse_code = store_info["warehouse_code"] if store_info and store_info.get("warehouse_code") else ""
        
        for row_idx, item in enumerate(sku_results, start=2):
            row_data = {
                "序号": 1,
                "门店编号": store_info["store_code"] if store_info else "",
                "门店三方编码": "",
                "仓库编码": warehouse_code,
                "加急程度\n（0：普通，1：加急）": self.DEFAULT_VALUES["加急程度"],
                "商品SKU编号": item["sku_code"],
                "商品三方SPEC编号": "",
                "单位类型": item["unit_type"],
                "出库数量": item["quantity"],
                "指定库存状态": self.DEFAULT_VALUES["指定库存状态"],
                "出库类型": self.DEFAULT_VALUES["出库类型"],
                "配送方式": self.DEFAULT_VALUES["配送方式"],
                "指定车型（专配）": "",
                "是否垫付": self.DEFAULT_VALUES["是否垫付"],
                "付款方式": "",
                "快递公司": "",
                "单价": "",
                "总金额": "",
                "是否制定批次": self.DEFAULT_VALUES["是否制定批次"],
                "批次号": "",
                "生产日期": "",
                "备注": item["remark"],
                "生产厂家编号": "",
                "门店收货地址编码": "",
                "三方单号": order_data["order_no"],
                "业务模式": "",
                "业务类型": "",
                "收货人": store_info["contact_person"] if store_info else "",
                "联系电话": store_info["phone"] if store_info else "",
                "收货地址": store_info["address"] if store_info else "",
                "C端快递公司": ""
            }
            
            for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
                value = row_data.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = cell_border
        
        wb.save(output_file)


# 全局配置
_global_config = {
    "shipper_id": None,
    "db_config": None,
    "output_dir": "./output"
}


def configure(shipper_id: str, db_config: Dict, output_dir: Optional[str] = None):
    """配置全局参数"""
    _global_config["shipper_id"] = shipper_id
    _global_config["db_config"] = db_config
    _global_config["output_dir"] = output_dir or "./output"


def convert_order_to_huading(order_file: str, output_file: str = None) -> Dict[str, Any]:
    """将客户订单转换为华鼎出库单模板"""
    if not _global_config["shipper_id"] or not _global_config["db_config"]:
        return {
            "success": False,
            "message": "请先调用 configure() 配置全局参数"
        }
    
    skill = OrderToHuadingTemplate(
        shipper_id=_global_config["shipper_id"],
        db_config=_global_config["db_config"],
        output_dir=_global_config["output_dir"]
    )
    
    return skill.execute(order_file, output_file)


if __name__ == "__main__":
    print("请先调用 configure() 配置参数")